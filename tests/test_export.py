from __future__ import annotations

from pathlib import Path
import json

from openpyxl import load_workbook

from timetable_scraper.export import export_rows, write_autofix_report
from timetable_scraper.models import NormalizedRow


def test_export_rows_apply_uniform_body_style_across_written_rows(tmp_path: Path) -> None:
    template_path = next(Path(".").glob("*.xlsx")).resolve()
    rows = [
        NormalizedRow(
            program="Demo Program",
            faculty="Demo Faculty",
            week_type="Обидва",
            day="Понеділок",
            start_time="08:00",
            end_time="09:20",
            subject=f"Предмет {index}",
            teacher="доц. Іваненко",
            lesson_type="лекція",
            room="101",
            groups="1",
            course="1",
            sheet_name="1 курс",
        )
        for index in range(8)
    ]
    exported_files, _, _ = export_rows(rows, [], template_path=template_path, output_dir=tmp_path / "out")
    workbook = load_workbook(exported_files[0])
    sheet = workbook.active

    early_cell = sheet["A3"]
    late_cell = sheet["A10"]

    assert early_cell.font.name == late_cell.font.name
    assert early_cell.font.sz == late_cell.font.sz
    assert early_cell.alignment.horizontal == late_cell.alignment.horizontal
    assert early_cell.alignment.vertical == late_cell.alignment.vertical
    assert early_cell.alignment.wrap_text == late_cell.alignment.wrap_text
    assert late_cell.font.name != "Calibri"
    assert late_cell.value == "Обидва"


def test_export_rows_clear_template_sample_rows_when_sheet_is_short(tmp_path: Path) -> None:
    template_path = next(Path(".").glob("*.xlsx")).resolve()
    rows = [
        NormalizedRow(
            program="Short Program",
            faculty="Short Faculty",
            week_type="Обидва",
            day="Вівторок",
            start_time="09:30",
            end_time="10:50",
            subject="Мікроекономіка",
            teacher="доц. Петренко",
            lesson_type="семінар",
            room="202",
            groups="2",
            course="2",
            sheet_name="1 курс",
        )
    ]
    exported_files, _, _ = export_rows(rows, [], template_path=template_path, output_dir=tmp_path / "out")
    workbook = load_workbook(exported_files[0])
    sheet = workbook.active

    assert sheet["A3"].value == "Обидва"
    assert sheet["E3"].value == "Мікроекономіка"
    assert sheet["A4"].value is None
    assert sheet["E4"].value is None


def test_manifest_serializes_provenance_fields(tmp_path: Path) -> None:
    template_path = next(Path(".").glob("*.xlsx")).resolve()
    rows = [
        NormalizedRow(
            program="Demo Program",
            faculty="Demo Faculty",
            week_type="Обидва",
            day="Понеділок",
            start_time="08:00",
            end_time="09:20",
            subject="Предмет",
            source_name="demo-source",
            source_kind="web_page",
            source_root_url="https://example.edu/schedule",
            asset_locator="https://docs.google.com/spreadsheets/d/demo/edit",
            autofix_actions=["week_type_defaulted", "subject_cleaned"],
        )
    ]
    _, manifest_path, _ = export_rows(rows, [], template_path=template_path, output_dir=tmp_path / "out")
    payload = json.loads(manifest_path.read_text(encoding="utf-8").splitlines()[0])
    assert payload["source_name"] == "demo-source"
    assert payload["source_root_url"] == "https://example.edu/schedule"
    assert payload["asset_locator"] == "https://docs.google.com/spreadsheets/d/demo/edit"
    assert payload["autofix_actions"] == ["week_type_defaulted", "subject_cleaned"]


def test_export_rows_preserve_literal_text_that_starts_with_equals(tmp_path: Path) -> None:
    template_path = next(Path(".").glob("*.xlsx")).resolve()
    rows = [
        NormalizedRow(
            program="Demo Program",
            faculty="Demo Faculty",
            week_type="Обидва",
            day="Понеділок",
            start_time="08:00",
            end_time="09:20",
            subject="=encoded-token",
            sheet_name="1 курс",
        )
    ]
    exported_files, _, _ = export_rows(rows, [], template_path=template_path, output_dir=tmp_path / "out")
    workbook = load_workbook(exported_files[0], data_only=False)
    sheet = workbook.active
    assert sheet["E3"].value == "=encoded-token"
    assert sheet["E3"].data_type == "s"


def test_write_autofix_report_creates_summary_and_row_sheets(tmp_path: Path) -> None:
    rows = [
        NormalizedRow(
            program="Demo Program",
            faculty="Demo Faculty",
            week_type="ÐžÐ±Ð¸Ð´Ð²Ð°",
            day="ÐŸÐ¾Ð½ÐµÐ´Ñ–Ð»Ð¾Ðº",
            start_time="08:00",
            end_time="09:20",
            subject="ÐŸÑ€ÐµÐ´Ð¼ÐµÑ‚",
            source_name="demo-source",
            source_root_url="https://example.edu/schedule",
            asset_locator="https://example.edu/file.xlsx",
            autofix_actions=["week_type_defaulted", "room_from_subject"],
            warnings=["subject_inferred_from_notes"],
        )
    ]

    json_path, xlsx_path, autofix_rows = write_autofix_report(rows, output_dir=tmp_path / "out")

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["rows_with_autofix"] == 1
    assert payload["action_counts"]["week_type_defaulted"] == 1
    assert payload["action_counts"]["room_from_subject"] == 1
    assert payload["rows"][0]["autofix_actions"] == ["week_type_defaulted", "room_from_subject"]

    workbook = load_workbook(xlsx_path)
    assert workbook.sheetnames == ["summary", "rows"]
    assert workbook["summary"]["A2"].value == "room_from_subject"
    assert workbook["rows"]["I2"].value == "week_type_defaulted, room_from_subject"
    assert autofix_rows == 1


def test_export_rows_truncate_overlong_program_filenames(tmp_path: Path) -> None:
    template_path = next(Path(".").glob("*.xlsx")).resolve()
    long_program = "Program " + ("very long title " * 20)
    rows = [
        NormalizedRow(
            program=long_program,
            faculty="Faculty " + ("name " * 10),
            week_type="ÐžÐ±Ð¸Ð´Ð²Ð°",
            day="ÐŸÐ¾Ð½ÐµÐ´Ñ–Ð»Ð¾Ðº",
            start_time="08:00",
            end_time="09:20",
            subject="ÐŸÑ€ÐµÐ´Ð¼ÐµÑ‚",
            sheet_name="1 ÐºÑƒÑ€Ñ",
        )
    ]

    exported_files, _, _ = export_rows(rows, [], template_path=template_path, output_dir=tmp_path / "out")

    assert len(exported_files) == 1
    assert exported_files[0].exists()
    assert len(exported_files[0].stem) <= 120
    assert len(exported_files[0].parent.name) <= 80


def test_export_rows_normalize_recovered_program_alias_for_filename(tmp_path: Path) -> None:
    template_path = next(Path(".").glob("*.xlsx")).resolve()
    rows = [
        NormalizedRow(
            program="",
            faculty="ННЦ Інститут біології та медицини",
            week_type="Обидва",
            day="Понеділок",
            start_time="08:30",
            end_time="09:50",
            subject="Біоінформатика",
            groups="Генетичнии аналіз",
            sheet_name="1 курс",
            asset_locator="fixtures/генетичнии аналіз.xlsx",
        )
    ]

    exported_files, _, _ = export_rows(rows, [], template_path=template_path, output_dir=tmp_path / "out")

    assert len(exported_files) == 1
    assert exported_files[0].stem == "Генетичний аналіз"


def test_export_rows_groups_same_program_by_course_sheets(tmp_path: Path) -> None:
    template_path = next(Path(".").glob("*.xlsx")).resolve()
    rows = [
        NormalizedRow(
            program="Соціологія",
            faculty="Факультет соціології",
            week_type="Верхній",
            day="Понеділок",
            start_time="08:30",
            end_time="09:50",
            subject="Соціологічна теорія",
            lesson_type="лекція",
            course="1",
            sheet_name="3 к 1с",
        ),
        NormalizedRow(
            program="Соціологія",
            faculty="Факультет соціології",
            week_type="Нижній",
            day="Вівторок",
            start_time="10:00",
            end_time="11:20",
            subject="Методи соціологічного дослідження",
            lesson_type="практичне заняття",
            course="2.0",
            sheet_name="2с 25 26",
        ),
    ]

    exported_files, _, _ = export_rows(rows, [], template_path=template_path, output_dir=tmp_path / "out")

    assert len(exported_files) == 1
    assert exported_files[0].stem == "Соціологія"
    workbook = load_workbook(exported_files[0])
    assert workbook.sheetnames == ["1 курс", "2 курс"]


def test_export_rows_do_not_use_compact_technical_sheet_names(tmp_path: Path) -> None:
    template_path = next(Path(".").glob("*.xlsx")).resolve()
    rows = [
        NormalizedRow(
            program="Соціологія",
            faculty="Факультет соціології",
            week_type="Верхній",
            day="Понеділок",
            start_time="08:30",
            end_time="09:50",
            subject="Соціологічна теорія",
            lesson_type="лекція",
            sheet_name="1к 1с 25-26",
        ),
        NormalizedRow(
            program="Соціологія",
            faculty="Факультет соціології",
            week_type="Нижній",
            day="Вівторок",
            start_time="10:00",
            end_time="11:20",
            subject="Методи соціологічного дослідження",
            lesson_type="практичне заняття",
            sheet_name="2с 25-26",
        ),
        NormalizedRow(
            program="Соціологія",
            faculty="Факультет соціології",
            week_type="Обидва",
            day="Середа",
            start_time="11:40",
            end_time="13:00",
            subject="Sociology of Gender",
            lesson_type="лекція",
            sheet_name="English 1c",
        ),
    ]

    exported_files, _, _ = export_rows(rows, [], template_path=template_path, output_dir=tmp_path / "out")

    assert len(exported_files) == 1
    workbook = load_workbook(exported_files[0])
    assert "1 курс" in workbook.sheetnames
    assert "Соціологія" in workbook.sheetnames
    assert "1к 1с 25-26" not in workbook.sheetnames
    assert "2с 25-26" not in workbook.sheetnames
    assert "English 1c" not in workbook.sheetnames
