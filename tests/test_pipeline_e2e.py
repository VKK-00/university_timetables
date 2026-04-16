from __future__ import annotations

from dataclasses import asdict
import json
from pathlib import Path

import requests
import pytest
from openpyxl import Workbook, load_workbook

from timetable_scraper.config import load_config
from timetable_scraper.models import AppConfig, DiscoveredAsset, DiscoveryResult, FetchedAsset, NormalizedRow, SourceConfig
from timetable_scraper.pipeline import run_pipeline, run_pipeline_batched


def test_pipeline_runs_against_real_archive(tmp_path: Path) -> None:
    config_path = Path("config/sources.yaml").resolve()
    config = load_config(config_path)
    if any(source.path and not source.path.exists() for source in config.sources):
        pytest.skip("Real archive fixture is not available in the working tree")
    config.output_dir = tmp_path / "out"
    config.cache_dir = tmp_path / "cache"
    result = run_pipeline(config)
    assert result.exported_files
    assert result.manifest_path.exists()
    assert result.review_queue_path.exists()
    assert result.review_summary_json_path and result.review_summary_json_path.exists()
    assert result.review_summary_xlsx_path and result.review_summary_xlsx_path.exists()
    assert result.run_delta_path and result.run_delta_path.exists()
    assert result.autofix_report_json_path and result.autofix_report_json_path.exists()
    assert result.autofix_report_xlsx_path and result.autofix_report_xlsx_path.exists()
    sample_workbook = load_workbook(result.exported_files[0])
    assert sample_workbook.sheetnames
    assert result.rows


def test_pipeline_skips_unavailable_assets(monkeypatch, tmp_path: Path) -> None:
    config = AppConfig(
        template_path=Path("Шаблон.xlsx").resolve(),
        output_dir=tmp_path / "out",
        cache_dir=tmp_path / "cache",
        confidence_threshold=0.74,
        ocr_enabled=False,
        sources=[],
    )
    bad_asset = DiscoveredAsset(
        source_name="bad",
        source_kind="google_sheet",
        source_url_or_path="https://docs.google.com/spreadsheets/d/bad/edit",
        asset_kind="google_sheet",
        locator="https://docs.google.com/spreadsheets/d/bad/edit",
        display_name="bad",
    )
    good_asset = DiscoveredAsset(
        source_name="good",
        source_kind="google_sheet",
        source_url_or_path="https://docs.google.com/spreadsheets/d/good/edit",
        asset_kind="google_sheet",
        locator="https://docs.google.com/spreadsheets/d/good/edit",
        display_name="good",
    )
    good_fetched = FetchedAsset(
        asset=good_asset,
        content=b"",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content_hash="good",
        resolved_locator="good.xlsx",
    )
    good_row = NormalizedRow(
        program="Algorithms",
        faculty="FIT",
        week_type="Обидва",
        day="Понеділок",
        start_time="09:30",
        end_time="10:50",
        subject="Theory of algorithms",
        groups="Algorithms",
        confidence=1.0,
        source_name="good",
        asset_locator=good_asset.locator,
    )

    def fake_fetch(asset, session, cache_dir):
        if asset.locator == bad_asset.locator:
            raise requests.HTTPError("401")
        return good_fetched

    def fake_export(rows, review_rows, *, template_path, output_dir):
        manifest_path = output_dir / "manifest.jsonl"
        review_path = output_dir / "review_queue.xlsx"
        output_dir.mkdir(parents=True, exist_ok=True)
        manifest_path.write_text("", encoding="utf-8")
        review_path.write_text("", encoding="utf-8")
        exported = output_dir / "demo.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Algorithms"
        sheet["A1"] = "Algorithms"
        sheet["A2"] = "Тиждень"
        sheet["B2"] = "День"
        sheet["C2"] = "Початок"
        sheet["D2"] = "Кінець"
        sheet["E2"] = "Назва предмета"
        sheet["A3"] = "Обидва"
        sheet["B3"] = "Понеділок"
        sheet["C3"] = "09:30"
        sheet["D3"] = "10:50"
        sheet["E3"] = "Theory of algorithms"
        workbook.save(exported)
        return [exported], manifest_path, review_path

    monkeypatch.setattr(
        "timetable_scraper.pipeline.discover_sources",
        lambda sources, session: DiscoveryResult(assets=[bad_asset, good_asset], issues=[]),
    )
    monkeypatch.setattr("timetable_scraper.pipeline.fetch_asset", fake_fetch)
    monkeypatch.setattr("timetable_scraper.pipeline.parse_asset", lambda fetched, ocr_enabled: fetched)
    monkeypatch.setattr("timetable_scraper.pipeline.normalize_document", lambda parsed: [good_row])
    monkeypatch.setattr("timetable_scraper.pipeline.export_rows", fake_export)

    result = run_pipeline(config)

    assert len(result.rows) == 1
    assert not result.review_rows
    assert result.exported_files
    assert result.review_summary_json_path and result.review_summary_json_path.exists()
    assert result.review_summary_xlsx_path and result.review_summary_xlsx_path.exists()
    assert result.run_delta_path and result.run_delta_path.exists()
    assert result.autofix_report_json_path and result.autofix_report_json_path.exists()
    assert result.autofix_report_xlsx_path and result.autofix_report_xlsx_path.exists()
    assert result.qa_failures == 0


def test_pipeline_cleans_previous_output_dir(tmp_path: Path, monkeypatch) -> None:
    config = AppConfig(
        template_path=Path("Шаблон.xlsx").resolve(),
        output_dir=tmp_path / "out",
        cache_dir=tmp_path / "cache",
        confidence_threshold=0.74,
        ocr_enabled=False,
        sources=[],
    )
    stale_file = config.output_dir / "stale.txt"
    stale_file.parent.mkdir(parents=True, exist_ok=True)
    stale_file.write_text("old", encoding="utf-8")

    monkeypatch.setattr(
        "timetable_scraper.pipeline.discover_sources",
        lambda sources, session: DiscoveryResult(assets=[], issues=[]),
    )
    monkeypatch.setattr(
        "timetable_scraper.pipeline.export_rows",
        lambda rows, review_rows, *, template_path, output_dir: ([], output_dir / "manifest.jsonl", output_dir / "review_queue.xlsx"),
    )
    monkeypatch.setattr(
        "timetable_scraper.pipeline.write_autofix_report",
        lambda rows, *, output_dir: (output_dir / "autofix_report.json", output_dir / "autofix_report.xlsx", 0),
    )
    monkeypatch.setattr(
        "timetable_scraper.pipeline.audit_exported_workbooks",
        lambda exported_files, output_dir: ([], output_dir / "qa_report.json", output_dir / "qa_report.xlsx"),
    )
    monkeypatch.setattr(
        "timetable_scraper.pipeline.write_source_summaries",
        lambda summaries, *, output_dir: (output_dir / "source_summary.json", output_dir / "source_summary.md"),
    )

    run_pipeline(config)

    assert not stale_file.exists()


def test_run_pipeline_batched_processes_sources_in_chunks(tmp_path: Path, monkeypatch) -> None:
    config = AppConfig(
        template_path=Path("Шаблон.xlsx").resolve(),
        output_dir=tmp_path / "out",
        cache_dir=tmp_path / "cache",
        confidence_threshold=0.74,
        ocr_enabled=False,
        sources=[
            SourceConfig(kind="web_page", name="history-schedule", url="https://example.test/history"),
            SourceConfig(kind="web_page", name="fit-schedule", url="https://example.test/fit"),
            SourceConfig(kind="web_page", name="phys-schedule", url="https://example.test/phys"),
        ],
    )
    batches: list[list[str]] = []

    def fake_discover_sources(sources, session):
        batches.append([source.name for source in sources])
        assets = [
            DiscoveredAsset(
                source_name=source.name,
                source_kind=source.kind,
                source_url_or_path=source.url or "",
                asset_kind="file_url",
                locator=f"{source.url}/asset.xlsx",
                display_name=source.name,
            )
            for source in sources
        ]
        return DiscoveryResult(assets=assets, issues=[])

    def fake_fetch(asset, session, cache_dir):
        return FetchedAsset(
            asset=asset,
            content=b"",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content_hash=asset.source_name,
            resolved_locator=f"{asset.source_name}.xlsx",
        )

    def fake_parse_asset(fetched, ocr_enabled):
        return fetched

    def fake_normalize_document(parsed):
        return [
            NormalizedRow(
                program="Applied Mathematics",
                faculty="Test faculty",
                week_type="Обидва",
                day="Понеділок",
                start_time="09:30",
                end_time="10:50",
                subject=f"{parsed.asset.source_name} subject",
                groups="AM-1",
                confidence=1.0,
                source_name=parsed.asset.source_name,
                source_kind=parsed.asset.source_kind,
                source_root_url=parsed.asset.source_root_url,
                asset_locator=parsed.asset.locator,
            )
        ]

    def fake_export(rows, review_rows, *, template_path, output_dir):
        manifest_path = output_dir / "manifest.jsonl"
        review_path = output_dir / "review_queue.xlsx"
        output_dir.mkdir(parents=True, exist_ok=True)
        manifest_path.write_text("", encoding="utf-8")
        review_path.write_text("", encoding="utf-8")
        exported = output_dir / "demo.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "demo"
        sheet["A1"] = "demo"
        sheet["A2"] = "Тиждень"
        sheet["B2"] = "День"
        sheet["C2"] = "Початок"
        sheet["D2"] = "Кінець"
        sheet["E2"] = "Назва предмета"
        for row_index, row in enumerate(rows, start=3):
            sheet.cell(row_index, 1).value = row.week_type
            sheet.cell(row_index, 2).value = row.day
            sheet.cell(row_index, 3).value = row.start_time
            sheet.cell(row_index, 4).value = row.end_time
            sheet.cell(row_index, 5).value = row.subject
        workbook.save(exported)
        return [exported], manifest_path, review_path

    monkeypatch.setattr("timetable_scraper.pipeline.discover_sources", fake_discover_sources)
    monkeypatch.setattr("timetable_scraper.pipeline.fetch_asset", fake_fetch)
    monkeypatch.setattr("timetable_scraper.pipeline.parse_asset", fake_parse_asset)
    monkeypatch.setattr("timetable_scraper.pipeline.normalize_document", fake_normalize_document)
    monkeypatch.setattr("timetable_scraper.pipeline.export_rows", fake_export)
    monkeypatch.setattr("timetable_scraper.pipeline.partition_rows", lambda rows, threshold: (rows, []))
    monkeypatch.setattr("timetable_scraper.pipeline.refine_group_quality", lambda accepted, review: (accepted, review))
    monkeypatch.setattr("timetable_scraper.pipeline.sanitize_export_rows", lambda accepted, review: (accepted, review))

    result = run_pipeline_batched(config, batch_size=2)

    assert batches == [["history-schedule", "fit-schedule"], ["phys-schedule"]]
    assert len(result.rows) == 3
    assert not result.review_rows
    assert result.exported_files


def test_run_pipeline_batched_merge_existing_preserves_unselected_sources(tmp_path: Path, monkeypatch) -> None:
    output_dir = tmp_path / "out"
    output_dir.mkdir(parents=True)
    old_row = NormalizedRow(
        program="History",
        faculty="History faculty",
        week_type="Обидва",
        day="Понеділок",
        start_time="09:00",
        end_time="10:20",
        subject="Old history subject",
        groups="H-1",
        confidence=1.0,
        source_name="history-schedule",
        source_kind="web_page",
        source_root_url="https://example.test/history",
        asset_locator="https://example.test/history/old.xlsx",
    )
    (output_dir / "manifest.jsonl").write_text(json.dumps(asdict(old_row), ensure_ascii=False) + "\n", encoding="utf-8")
    (output_dir / "source_summary.json").write_text(
        json.dumps(
            [
                {
                    "source_name": "history-schedule",
                    "source_root_url": "https://example.test/history",
                    "status": "parsed",
                    "accepted_rows": 1,
                    "review_rows": 0,
                    "autofix_rows": 0,
                    "discovered_assets": 2,
                    "attempted_assets": 2,
                    "discovery_issues": [],
                    "runtime_issues": [],
                    "top_review_warnings": [],
                    "top_review_qa_flags": [],
                    "note": "old history metadata",
                },
                {
                    "source_name": "blocked-schedule",
                    "source_root_url": "https://example.test/blocked",
                    "status": "blocked",
                    "accepted_rows": 0,
                    "review_rows": 0,
                    "autofix_rows": 0,
                    "discovered_assets": 0,
                    "attempted_assets": 0,
                    "discovery_issues": ["Cloudflare"],
                    "runtime_issues": ["Cloudflare"],
                    "top_review_warnings": [],
                    "top_review_qa_flags": [],
                    "note": "old blocker metadata",
                },
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    fit_source = SourceConfig(kind="web_page", name="fit-schedule", url="https://example.test/fit")
    all_sources = [
        SourceConfig(kind="web_page", name="history-schedule", url="https://example.test/history"),
        SourceConfig(kind="web_page", name="blocked-schedule", url="https://example.test/blocked"),
        fit_source,
    ]
    config = AppConfig(
        template_path=Path("Шаблон.xlsx").resolve(),
        output_dir=output_dir,
        cache_dir=tmp_path / "cache",
        confidence_threshold=0.74,
        ocr_enabled=False,
        sources=[fit_source],
    )

    def fake_discover_sources(sources, session):
        return DiscoveryResult(
            assets=[
                DiscoveredAsset(
                    source_name="fit-schedule",
                    source_kind="web_page",
                    source_url_or_path="https://example.test/fit",
                    asset_kind="file_url",
                    locator="https://example.test/fit/asset.xlsx",
                    display_name="fit",
                )
            ],
            issues=[],
        )

    def fake_fetch(asset, session, cache_dir):
        return FetchedAsset(
            asset=asset,
            content=b"",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content_hash="fit",
            resolved_locator="fit.xlsx",
        )

    def fake_normalize_document(parsed):
        return [
            NormalizedRow(
                program="FIT",
                faculty="FIT faculty",
                week_type="Обидва",
                day="Вівторок",
                start_time="10:30",
                end_time="11:50",
                subject="New fit subject",
                groups="FIT-1",
                confidence=1.0,
                source_name="fit-schedule",
                source_kind="web_page",
                source_root_url="https://example.test/fit",
                asset_locator="https://example.test/fit/asset.xlsx",
            )
        ]

    def fake_export(rows, review_rows, *, template_path, output_dir):
        manifest_path = output_dir / "manifest.jsonl"
        review_path = output_dir / "review_queue.xlsx"
        output_dir.mkdir(parents=True, exist_ok=True)
        manifest_path.write_text(
            "\n".join(json.dumps(asdict(row), ensure_ascii=False) for row in rows + review_rows) + "\n",
            encoding="utf-8",
        )
        review_path.write_text("", encoding="utf-8")
        exported = output_dir / "demo.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "demo"
        sheet["A1"] = "demo"
        sheet["A2"] = "Тиждень"
        sheet["B2"] = "День"
        sheet["C2"] = "Початок"
        sheet["D2"] = "Кінець"
        sheet["E2"] = "Назва предмета"
        for row_index, row in enumerate(rows, start=3):
            sheet.cell(row_index, 1).value = row.week_type
            sheet.cell(row_index, 2).value = row.day
            sheet.cell(row_index, 3).value = row.start_time
            sheet.cell(row_index, 4).value = row.end_time
            sheet.cell(row_index, 5).value = row.subject
        workbook.save(exported)
        return [exported], manifest_path, review_path

    monkeypatch.setattr("timetable_scraper.pipeline.discover_sources", fake_discover_sources)
    monkeypatch.setattr("timetable_scraper.pipeline.fetch_asset", fake_fetch)
    monkeypatch.setattr("timetable_scraper.pipeline.parse_asset", lambda fetched, ocr_enabled: fetched)
    monkeypatch.setattr("timetable_scraper.pipeline.normalize_document", fake_normalize_document)
    monkeypatch.setattr("timetable_scraper.pipeline.export_rows", fake_export)
    monkeypatch.setattr("timetable_scraper.pipeline.partition_rows", lambda rows, threshold: (rows, []))
    monkeypatch.setattr("timetable_scraper.pipeline.refine_group_quality", lambda accepted, review: (accepted, review))
    monkeypatch.setattr("timetable_scraper.pipeline.sanitize_export_rows", lambda accepted, review: (accepted, review))

    result = run_pipeline_batched(config, batch_size=1, merge_existing=True, summary_sources=all_sources)

    assert {row.subject for row in result.rows} == {"Old history subject", "New fit subject"}
    manifest_sources = {
        json.loads(line)["source_name"]
        for line in result.manifest_path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    }
    assert manifest_sources == {"history-schedule", "fit-schedule"}
    summary_by_source = {summary.source_name: summary for summary in result.source_summaries}
    assert summary_by_source["history-schedule"].accepted_rows == 1
    assert summary_by_source["history-schedule"].discovered_assets == 2
    assert summary_by_source["blocked-schedule"].status == "blocked"
    assert summary_by_source["blocked-schedule"].runtime_issues == ["Cloudflare"]
    assert summary_by_source["fit-schedule"].accepted_rows == 1
