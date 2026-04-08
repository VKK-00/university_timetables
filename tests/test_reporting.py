from __future__ import annotations

from collections import Counter
from pathlib import Path
import json

from openpyxl import load_workbook

from timetable_scraper.models import DiscoveryResult, NormalizedRow, SourceConfig, SourceRunSummary
from timetable_scraper.reporting import (
    build_source_summaries,
    load_previous_source_summaries,
    write_review_summary,
    write_run_delta,
    write_source_summaries,
)


def test_write_review_summary_aggregates_warnings_flags_and_examples(tmp_path: Path) -> None:
    rows = [
        NormalizedRow(
            program="Physics",
            faculty="Physics",
            week_type="Обидва",
            day="",
            start_time="08:40",
            end_time="10:15",
            subject="",
            source_name="phys-schedule",
            source_root_url="https://phys.knu.ua/navchannya/rozklad-zanyat?ad",
            warnings=["missing_subject"],
            qa_flags=["missing_subject", "implausible_time"],
            raw_excerpt="пр 301",
        ),
        NormalizedRow(
            program="Physics",
            faculty="Physics",
            week_type="Обидва",
            day="",
            start_time="08:40",
            end_time="10:15",
            subject="",
            source_name="phys-schedule",
            source_root_url="https://phys.knu.ua/navchannya/rozklad-zanyat?ad",
            warnings=["missing_subject"],
            qa_flags=["missing_subject"],
            raw_excerpt="пр 301",
        ),
        NormalizedRow(
            program="FIT",
            faculty="FIT",
            week_type="Обидва",
            day="Понеділок",
            start_time="09:30",
            end_time="10:50",
            subject="",
            source_name="fit-schedule",
            source_root_url="https://fit.knu.ua/for-students/lessons-schedule",
            warnings=["garbage_text_subject"],
            qa_flags=["garbage_text", "inconsistent_columns"],
            raw_excerpt="ФІТ у 2023 році",
        ),
    ]

    json_path, xlsx_path = write_review_summary(rows, output_dir=tmp_path / "out")

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["total_review_rows"] == 3
    assert payload["source_count"] == 2
    assert payload["sources"][0]["source_name"] == "phys-schedule"
    assert payload["sources"][0]["warning_counts"]["missing_subject"] == 2
    assert payload["sources"][0]["qa_flag_counts"]["missing_subject"] == 2
    assert payload["sources"][0]["issues"][0]["top_raw_excerpts"][0]["excerpt"] == "пр 301"

    workbook = load_workbook(xlsx_path)
    assert workbook.sheetnames == ["summary", "issues"]
    assert workbook["summary"]["A2"].value == "phys-schedule"
    assert workbook["issues"]["B2"].value in {"warning", "qa_flag"}


def test_source_summary_and_run_delta_include_autofix_counts(tmp_path: Path) -> None:
    sources = [
        SourceConfig(kind="web_page", name="phys-schedule", url="https://phys.knu.ua/navchannya/rozklad-zanyat?ad"),
        SourceConfig(kind="web_page", name="philology-schedule", url="https://philology.knu.ua/nauka/aspirantura/rozklad-zanyat/"),
    ]
    accepted_rows = [
        NormalizedRow(
            program="Physics",
            faculty="Physics",
            week_type="Обидва",
            day="Понеділок",
            start_time="08:40",
            end_time="10:15",
            subject="Оптика",
            source_name="phys-schedule",
            source_root_url="https://phys.knu.ua/navchannya/rozklad-zanyat?ad",
            autofix_actions=["week_type_defaulted"],
        )
    ]
    review_rows = [
        NormalizedRow(
            program="Philology",
            faculty="Philology",
            week_type="Обидва",
            day="",
            start_time="08:40",
            end_time="10:15",
            subject="нечитаємий фрагмент",
            source_name="philology-schedule",
            source_root_url="https://philology.knu.ua/nauka/aspirantura/rozklad-zanyat/",
            warnings=["missing_day"],
            qa_flags=["garbage_text", "inconsistent_columns"],
        )
    ]
    summaries = build_source_summaries(
        sources,
        DiscoveryResult(assets=[], issues=[]),
        accepted_rows,
        review_rows,
        attempted_assets=Counter({"phys-schedule": 1, "philology-schedule": 1}),
        runtime_issues={},
    )

    source_summary_json, _ = write_source_summaries(summaries, output_dir=tmp_path / "out")
    previous = load_previous_source_summaries(tmp_path / "out")
    delta_path = write_run_delta(summaries, previous, output_dir=tmp_path / "delta")

    summary_payload = json.loads(source_summary_json.read_text(encoding="utf-8"))
    assert summary_payload[0]["autofix_rows"] == 1
    assert summary_payload[1]["status"] in {"review-only", "confirmed-blocker"}

    delta_payload = json.loads(delta_path.read_text(encoding="utf-8"))
    assert delta_payload["previous_run_found"] is True
    phys_item = next(item for item in delta_payload["sources"] if item["source_name"] == "phys-schedule")
    assert phys_item["delta"]["accepted_rows"]["current"] == 1
    assert phys_item["delta"]["autofix_rows"]["current"] == 1
    assert phys_item["status_current"] == "parsed"


def test_load_previous_source_summaries_handles_missing_file(tmp_path: Path) -> None:
    assert load_previous_source_summaries(tmp_path / "missing") == {}


def test_run_delta_uses_previous_counts_when_present(tmp_path: Path) -> None:
    previous_payload = [
        {
            "source_name": "fit-schedule",
            "source_root_url": "https://fit.knu.ua/for-students/lessons-schedule",
            "status": "parsed",
            "accepted_rows": 10,
            "review_rows": 5,
            "autofix_rows": 3,
        }
    ]
    summary_path = tmp_path / "out" / "source_summary.json"
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    summary_path.write_text(json.dumps(previous_payload, ensure_ascii=False), encoding="utf-8")

    current = [
        SourceRunSummary(
            source_name="fit-schedule",
            source_root_url="https://fit.knu.ua/for-students/lessons-schedule",
            status="parsed",
            accepted_rows=14,
            review_rows=4,
            autofix_rows=6,
        )
    ]

    previous = load_previous_source_summaries(tmp_path / "out")
    delta_path = write_run_delta(current, previous, output_dir=tmp_path / "out")
    payload = json.loads(delta_path.read_text(encoding="utf-8"))
    assert payload["totals"]["accepted_rows"]["delta"] == 4
    assert payload["totals"]["review_rows"]["delta"] == -1
    assert payload["totals"]["autofix_rows"]["delta"] == 3


def test_source_summary_marks_garbled_review_only_source_as_confirmed_blocker() -> None:
    sources = [
        SourceConfig(
            kind="web_page",
            name="philology-schedule",
            url="https://philology.knu.ua/nauka/aspirantura/rozklad-zanyat/",
        )
    ]
    review_rows = [
        NormalizedRow(
            program="Philology",
            faculty="Philology",
            week_type="Обидва",
            day="",
            start_time="10:00",
            end_time="11:20",
            subject="garbled fragment",
            source_name="philology-schedule",
            source_root_url="https://philology.knu.ua/nauka/aspirantura/rozklad-zanyat/",
            warnings=["missing_subject", "garbage_text_subject"],
            qa_flags=["garbage_text", "inconsistent_columns", "subject_too_long"],
        )
    ]

    summaries = build_source_summaries(
        sources,
        DiscoveryResult(assets=[], issues=[]),
        [],
        review_rows,
        attempted_assets=Counter({"philology-schedule": 1}),
        runtime_issues={},
    )

    assert summaries[0].status == "confirmed-blocker"
    assert "too noisy" in summaries[0].note


def test_source_summary_marks_http_410_asset_as_confirmed_blocker() -> None:
    sources = [
        SourceConfig(
            kind="web_page",
            name="csc-schedule",
            url="https://www.csc.knu.ua/en/schedule",
        )
    ]

    summaries = build_source_summaries(
        sources,
        DiscoveryResult(assets=[], issues=[]),
        [],
        [],
        attempted_assets=Counter({"csc-schedule": 1}),
        runtime_issues={"csc-schedule": ["HTTPError: 410 Client Error: Gone for url: https://docs.google.com/spreadsheets/d/test/edit"]},
    )

    assert summaries[0].status == "confirmed-blocker"
    assert "HTTP 410" in summaries[0].note
