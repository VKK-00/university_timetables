from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from timetable_scraper.adapters.excel import parse_excel_asset
from timetable_scraper.models import DiscoveredAsset, FetchedAsset

WORKBOOKS_DIR = Path(__file__).parent / "fixtures" / "workbooks"


def _make_fetched(path: Path) -> FetchedAsset:
    asset = DiscoveredAsset(
        source_name="fixture",
        source_kind="zip",
        source_url_or_path="fixture.zip",
        asset_kind="zip_entry",
        locator=f"fixture.zip::{path.name}",
        display_name=path.name,
    )
    return FetchedAsset(
        asset=asset,
        content=path.read_bytes(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content_hash=path.stem,
        resolved_locator=path.name,
    )


@pytest.mark.parametrize(
    "filename",
    ["veb.xlsx", "physics_2.xlsx", "mev.xlsx", "mk.xlsx", "pravo.xlsx", "econ.xlsx"],
)
def test_fixture_workbooks_parse(filename: str) -> None:
    document = parse_excel_asset(_make_fetched(WORKBOOKS_DIR / filename))
    total_records = sum(len(sheet.records) for sheet in document.sheets)
    assert document.sheets
    assert total_records > 0


def test_physics_fixture_supports_subjectu_header_variant() -> None:
    document = parse_excel_asset(_make_fetched(WORKBOOKS_DIR / "physics_2.xlsx"))
    first_record = document.sheets[0].records[0]
    assert first_record.values["subject"] == "Електродинаміка"


def test_pravo_fixture_skips_empty_sheet_with_warning() -> None:
    document = parse_excel_asset(_make_fetched(WORKBOOKS_DIR / "pravo.xlsx"))
    assert any("Skipped empty sheet" in warning for warning in document.warnings)
    assert any(sheet.sheet_name == "1 курс" for sheet in document.sheets)


def test_fit_style_grid_workbook_is_parsed() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ІПЗ"
    worksheet["A2"] = "день"
    worksheet["B2"] = "час"
    worksheet["C2"] = "1 курс ІПЗ"
    worksheet.merge_cells("C2:D2")
    worksheet["C3"] = "група ІПЗ-11"
    worksheet.merge_cells("C3:D3")
    worksheet["C4"] = "підгрупа ІПЗ-11/1"
    worksheet.merge_cells("C4:D4")
    worksheet["A5"] = "понеділок"
    worksheet["B5"] = "9:00-10:20"
    worksheet["C5"] = "Архітектура комп'ютера (лаб) 12т"
    worksheet.merge_cells("C5:D7")
    worksheet["C8"] = "I тиждень"
    worksheet.merge_cells("C8:D8")
    worksheet["C9"] = "Вовна О. В."
    worksheet.merge_cells("C9:D9")
    worksheet["C10"] = "109 ауд."
    worksheet["D10"] = "meet"

    buffer = BytesIO()
    workbook.save(buffer)
    asset = DiscoveredAsset(
        source_name="fit-grid",
        source_kind="google_sheet",
        source_url_or_path="https://fit.knu.ua/for-students/lessons-schedule",
        asset_kind="google_sheet",
        locator="https://docs.google.com/spreadsheets/d/test/edit#gid=0",
        display_name="fit-grid.xlsx",
    )
    fetched = FetchedAsset(
        asset=asset,
        content=buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content_hash="fit-grid",
        resolved_locator="fit-grid.xlsx",
    )

    document = parse_excel_asset(fetched)
    total_records = sum(len(sheet.records) for sheet in document.sheets)

    assert total_records == 1
    record = document.sheets[0].records[0]
    assert record.values["day"] == "Понеділок"
    assert record.values["start_time"] == "09:00"
    assert record.values["end_time"] == "10:20"
    assert record.values["subject"] == "Архітектура комп'ютера"
    assert record.values["lesson_type"] == "лабораторна"
    assert record.values["week_type"] == "Верхній"
    assert record.values["teacher"] == "Вовна О. В."
    assert record.values["room"] == "109 ауд."


def test_generic_grid_workbook_is_parsed() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "1 курс"
    worksheet["C1"] = "1 курс"
    worksheet["C2"] = "група 1"
    worksheet["A4"] = "Понеділок"
    worksheet["B4"] = "8:40-10:15"
    worksheet["C4"] = "Алгебра (лек)"
    worksheet["C5"] = "доц. Іваненко І.І."
    worksheet["C6"] = "ауд. 101"
    worksheet["A8"] = "Вівторок"
    worksheet["B8"] = "10:35-12:10"
    worksheet["C8"] = "Геометрія"

    buffer = BytesIO()
    workbook.save(buffer)
    asset = DiscoveredAsset(
        source_name="generic-grid",
        source_kind="file_url",
        source_url_or_path="https://example.edu/schedule.xlsx",
        asset_kind="file_url",
        locator="https://example.edu/schedule.xlsx",
        display_name="schedule.xlsx",
    )
    fetched = FetchedAsset(
        asset=asset,
        content=buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content_hash="generic-grid",
        resolved_locator="generic-grid.xlsx",
    )

    document = parse_excel_asset(fetched)
    total_records = sum(len(sheet.records) for sheet in document.sheets)

    assert total_records >= 1
    record = document.sheets[0].records[0]
    assert record.values["day"] == "Понеділок"
    assert record.values["start_time"] == "08:40"
    assert record.values["end_time"] == "10:15"
    assert "Алгебра" in record.values["subject"]


def test_generic_grid_workbook_supports_days_with_dates() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Журналістика"
    worksheet["C1"] = "1 курс"
    worksheet["C2"] = "група 1"
    worksheet["A4"] = "Понеділок (02.09.2019)"
    worksheet["B4"] = "14:10"
    worksheet["C4"] = "Вступ до журналістики"
    worksheet["C5"] = "доц. Іваненко І.І."
    worksheet["A6"] = "Вівторок (03.09.2019)"
    worksheet["B6"] = "15:50"
    worksheet["C6"] = "Медіаправо"
    worksheet["A8"] = "Середа (04.09.2019)"
    worksheet["B8"] = "17:30"
    worksheet["C8"] = "Практикум"

    buffer = BytesIO()
    workbook.save(buffer)
    asset = DiscoveredAsset(
        source_name="generic-grid-dates",
        source_kind="file_url",
        source_url_or_path="https://example.edu/journ.xlsx",
        asset_kind="file_url",
        locator="https://example.edu/journ.xlsx",
        display_name="journ.xlsx",
    )
    fetched = FetchedAsset(
        asset=asset,
        content=buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content_hash="generic-grid-dates",
        resolved_locator="journ.xlsx",
    )

    document = parse_excel_asset(fetched)
    records = [record for sheet in document.sheets for record in sheet.records]

    assert records
    assert records[0].values["day"] == "Понеділок"
    assert records[0].values["start_time"] == "14:10"


def test_generic_grid_workbook_skips_mid_sheet_headers_and_room_only_cells() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "\u041b\u0438\u0441\u04421"
    worksheet["C1"] = "\u0420\u043e\u0437\u043a\u043b\u0430\u0434 \u0437\u0430\u043d\u044f\u0442\u044c 2025 2026 \u043d\u0430\u0432\u0447.\u0440."
    worksheet["C2"] = "1 \u043a\u0443\u0440\u0441"
    worksheet["C3"] = "\u0433\u0440\u0443\u043f\u0430 1"
    worksheet["A4"] = "\u041f\u043e\u043d\u0435\u0434\u0456\u043b\u043e\u043a"
    worksheet["B4"] = "8:40-10:15"
    worksheet["C4"] = "\u0410\u043b\u0433\u0435\u0431\u0440\u0430"
    worksheet["C5"] = "\u0434\u043e\u0446. \u0406\u0432\u0430\u043d\u0435\u043d\u043a\u043e \u0406.\u0406."
    worksheet["A8"] = "I\u0406 \u043a\u0443\u0440\u0441"
    worksheet["C8"] = "\u0424\u0443\u043d\u0434\u0430\u043c\u0435\u043d\u0442\u0430\u043b\u044c\u043d\u0430 \u043c\u0435\u0434\u0438\u0447\u043d\u0430 \u0444\u0456\u0437\u0438\u043a\u0430"
    worksheet["A9"] = "\u0412\u0456\u0432\u0442\u043e\u0440\u043e\u043a"
    worksheet["B9"] = "8:40-10:15"
    worksheet["C9"] = "\u043f\u0440. 301"
    worksheet["C10"] = "\u041b\u0456\u043d\u0456\u0439\u043d\u0430 \u0430\u043b\u0433\u0435\u0431\u0440\u0430"
    worksheet["C11"] = "\u0434\u043e\u0446. \u041f\u0435\u0442\u0440\u043e\u0432 \u041f.\u041f."

    buffer = BytesIO()
    workbook.save(buffer)
    asset = DiscoveredAsset(
        source_name="generic-grid",
        source_kind="file_url",
        source_url_or_path="https://example.edu/schedule.xlsx",
        asset_kind="file_url",
        locator="https://example.edu/schedule.xlsx",
        display_name="schedule.xlsx",
    )
    fetched = FetchedAsset(
        asset=asset,
        content=buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content_hash="generic-grid-headers",
        resolved_locator="generic-grid-headers.xlsx",
    )

    document = parse_excel_asset(fetched)
    records = [record for sheet in document.sheets for record in sheet.records]
    subjects = {record.values["subject"] for record in records}

    assert "\u0410\u043b\u0433\u0435\u0431\u0440\u0430" in subjects
    assert "\u041b\u0456\u043d\u0456\u0439\u043d\u0430 \u0430\u043b\u0433\u0435\u0431\u0440\u0430" in subjects
    assert "\u0424\u0443\u043d\u0434\u0430\u043c\u0435\u043d\u0442\u0430\u043b\u044c\u043d\u0430 \u043c\u0435\u0434\u0438\u0447\u043d\u0430 \u0444\u0456\u0437\u0438\u043a\u0430" not in subjects
    assert "\u043f\u0440. 301" not in subjects
