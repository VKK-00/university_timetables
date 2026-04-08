from __future__ import annotations

from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

from .models import NormalizedRow
from .utils import coalesce_label, ensure_parent, json_dumps, slugify_filename, truncate_sheet_title


REVIEW_COLUMNS = [
    "program",
    "faculty",
    "sheet_name",
    "confidence",
    "warnings",
    "autofix_actions",
    "qa_flags",
    "qa_severity",
    "source_name",
    "source_kind",
    "source_root_url",
    "asset_locator",
    "raw_excerpt",
    "week_type",
    "week_source",
    "day",
    "start_time",
    "end_time",
    "subject",
    "teacher",
    "lesson_type",
    "link",
    "room",
    "groups",
    "course",
    "notes",
]

BODY_COLUMNS = [
    "week_type",
    "day",
    "start_time",
    "end_time",
    "subject",
    "teacher",
    "lesson_type",
    "link",
    "room",
    "groups",
    "course",
    "notes",
]

MAX_EXPORT_FACULTY_LENGTH = 80
MAX_EXPORT_PROGRAM_LENGTH = 120


@dataclass(slots=True)
class CellStyleSnapshot:
    font: object
    fill: object
    border: object
    alignment: object
    protection: object
    number_format: str


@dataclass(slots=True)
class TemplateStylePack:
    title_style: CellStyleSnapshot
    header_style: CellStyleSnapshot
    body_style: CellStyleSnapshot
    title_row_height: float | None
    header_row_height: float | None
    body_row_height: float | None


def export_rows(
    rows: list[NormalizedRow],
    review_rows: list[NormalizedRow],
    *,
    template_path: Path,
    output_dir: Path,
) -> tuple[list[Path], Path, Path]:
    exported_files = _export_program_workbooks(rows, template_path=template_path, output_dir=output_dir)
    manifest_path = output_dir / "manifest.jsonl"
    review_queue_path = output_dir / "review_queue.xlsx"
    _write_manifest([*rows, *review_rows], manifest_path)
    _write_review_queue(review_rows, review_queue_path, template_path=template_path)
    return exported_files, manifest_path, review_queue_path


def write_autofix_report(rows: list[NormalizedRow], *, output_dir: Path) -> tuple[Path, Path, int]:
    autofixed_rows = [row for row in rows if row.autofix_actions]
    action_counts: Counter[str] = Counter()
    for row in autofixed_rows:
        action_counts.update(row.autofix_actions)

    json_path = output_dir / "autofix_report.json"
    xlsx_path = output_dir / "autofix_report.xlsx"
    _write_autofix_report_json(autofixed_rows, action_counts, json_path)
    _write_autofix_report_xlsx(autofixed_rows, action_counts, xlsx_path)
    return json_path, xlsx_path, len(autofixed_rows)


def _export_program_workbooks(rows: list[NormalizedRow], *, template_path: Path, output_dir: Path) -> list[Path]:
    grouped: dict[tuple[str, str], list[NormalizedRow]] = defaultdict(list)
    for row in rows:
        grouped[(_export_faculty_label(row), _export_program_label(row))].append(row)

    exported: list[Path] = []
    for (faculty, program), program_rows in grouped.items():
        workbook = load_workbook(template_path)
        template_sheet = workbook.active
        style_pack = _capture_template_styles(template_sheet)
        _prepare_output_sheet(template_sheet, program=program, style_pack=style_pack)
        sheet_map: dict[str, list[NormalizedRow]] = defaultdict(list)
        for row in sorted(program_rows, key=_sort_key):
            sheet_map[_export_sheet_label(row)].append(row)
        first_sheet = True
        for sheet_name, sheet_rows in sheet_map.items():
            sheet = template_sheet if first_sheet else workbook.copy_worksheet(template_sheet)
            first_sheet = False
            sheet.title = truncate_sheet_title(sheet_name)
            if sheet is not template_sheet:
                _prepare_output_sheet(sheet, program=program, style_pack=style_pack)
            for row_index, row in enumerate(sheet_rows, start=3):
                _write_body_row(sheet, row_index, row, style_pack)
        faculty_dir = _truncate_export_name(slugify_filename(faculty), max_length=MAX_EXPORT_FACULTY_LENGTH)
        program_name = _truncate_export_name(slugify_filename(program), max_length=MAX_EXPORT_PROGRAM_LENGTH)
        target = output_dir / faculty_dir / f"{program_name}.xlsx"
        ensure_parent(target)
        workbook.save(target)
        exported.append(target)
    return exported


def _write_manifest(rows: Iterable[NormalizedRow], path: Path) -> None:
    ensure_parent(path)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(
                json_dumps(
                    {
                        "program": row.program,
                        "faculty": row.faculty,
                        "sheet_name": row.sheet_name,
                        "week_type": row.week_type,
                        "week_source": row.week_source,
                        "day": row.day,
                        "start_time": row.start_time,
                        "end_time": row.end_time,
                        "subject": row.subject,
                        "teacher": row.teacher,
                        "lesson_type": row.lesson_type,
                        "link": row.link,
                        "room": row.room,
                        "groups": row.groups,
                        "course": row.course,
                        "notes": row.notes,
                        "source_name": row.source_name,
                        "source_kind": row.source_kind,
                        "source_root_url": row.source_root_url,
                        "asset_locator": row.asset_locator,
                        "source_url_or_path": row.source_root_url,
                        "confidence": row.confidence,
                        "warnings": row.warnings,
                        "autofix_actions": row.autofix_actions,
                        "qa_flags": row.qa_flags,
                        "qa_severity": row.qa_severity,
                        "raw_excerpt": row.raw_excerpt,
                        "content_hash": row.content_hash,
                    }
                )
                + "\n"
            )


def _write_review_queue(rows: list[NormalizedRow], path: Path, *, template_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "review_queue"
    style_pack = _capture_template_styles(load_workbook(template_path).active)
    for column, title in enumerate(REVIEW_COLUMNS, start=1):
        cell = sheet.cell(1, column)
        cell.value = title
        _apply_cell_style(cell, style_pack.header_style)
    if style_pack.header_row_height is not None:
        sheet.row_dimensions[1].height = style_pack.header_row_height
    for row_index, row in enumerate(rows, start=2):
        values = {
            "program": row.program,
            "faculty": row.faculty,
            "sheet_name": row.sheet_name,
            "confidence": row.confidence,
            "warnings": ", ".join(row.warnings),
            "autofix_actions": ", ".join(row.autofix_actions),
            "qa_flags": ", ".join(row.qa_flags),
            "qa_severity": row.qa_severity,
            "source_name": row.source_name,
            "source_kind": row.source_kind,
            "source_root_url": row.source_root_url,
            "asset_locator": row.asset_locator,
            "raw_excerpt": row.raw_excerpt,
            "week_type": row.week_type,
            "week_source": row.week_source,
            "day": row.day,
            "start_time": row.start_time,
            "end_time": row.end_time,
            "subject": row.subject,
            "teacher": row.teacher,
            "lesson_type": row.lesson_type,
            "link": row.link,
            "room": row.room,
            "groups": row.groups,
            "course": row.course,
            "notes": row.notes,
        }
        for column, title in enumerate(REVIEW_COLUMNS, start=1):
            cell = sheet.cell(row_index, column)
            _set_cell_value(cell, values.get(title, ""))
            _apply_cell_style(cell, style_pack.body_style)
        if style_pack.body_row_height is not None:
            sheet.row_dimensions[row_index].height = style_pack.body_row_height
    ensure_parent(path)
    workbook.save(path)


def _prepare_output_sheet(sheet, *, program: str, style_pack: TemplateStylePack) -> None:
    sheet["A1"] = program
    _apply_cell_style(sheet["A1"], style_pack.title_style)
    if style_pack.title_row_height is not None:
        sheet.row_dimensions[1].height = style_pack.title_row_height
    if style_pack.header_row_height is not None:
        sheet.row_dimensions[2].height = style_pack.header_row_height
    for row_index in range(3, sheet.max_row + 1):
        for column in range(1, len(BODY_COLUMNS) + 1):
            sheet.cell(row_index, column).value = None


def _write_autofix_report_json(rows: list[NormalizedRow], action_counts: Counter[str], path: Path) -> None:
    ensure_parent(path)
    payload = {
        "rows_with_autofix": len(rows),
        "action_counts": dict(sorted(action_counts.items())),
        "rows": [
            {
                "faculty": row.faculty,
                "program": row.program,
                "sheet_name": row.sheet_name,
                "subject": row.subject,
                "source_name": row.source_name,
                "source_root_url": row.source_root_url,
                "asset_locator": row.asset_locator,
                "qa_severity": row.qa_severity,
                "autofix_actions": row.autofix_actions,
                "warnings": row.warnings,
                "raw_excerpt": row.raw_excerpt,
            }
            for row in rows
        ],
    }
    path.write_text(json_dumps(payload) + "\n", encoding="utf-8")


def _write_autofix_report_xlsx(rows: list[NormalizedRow], action_counts: Counter[str], path: Path) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_headers = ["action", "count"]
    for column, title in enumerate(summary_headers, start=1):
        summary_sheet.cell(1, column).value = title
    row_index = 2
    for action, count in sorted(action_counts.items()):
        summary_sheet.cell(row_index, 1).value = action
        summary_sheet.cell(row_index, 2).value = count
        row_index += 1

    rows_sheet = workbook.create_sheet("rows")
    row_headers = [
        "faculty",
        "program",
        "sheet_name",
        "subject",
        "source_name",
        "source_root_url",
        "asset_locator",
        "qa_severity",
        "autofix_actions",
        "warnings",
        "raw_excerpt",
    ]
    for column, title in enumerate(row_headers, start=1):
        rows_sheet.cell(1, column).value = title
    for row_index, row in enumerate(rows, start=2):
        values = [
            row.faculty,
            row.program,
            row.sheet_name,
            row.subject,
            row.source_name,
            row.source_root_url,
            row.asset_locator,
            row.qa_severity,
            ", ".join(row.autofix_actions),
            ", ".join(row.warnings),
            row.raw_excerpt,
        ]
        for column, value in enumerate(values, start=1):
            _set_cell_value(rows_sheet.cell(row_index, column), value)
    ensure_parent(path)
    workbook.save(path)


def _write_body_row(sheet, row_index: int, row: NormalizedRow, style_pack: TemplateStylePack) -> None:
    values = [
        row.week_type,
        row.day,
        row.start_time,
        row.end_time,
        row.subject,
        row.teacher,
        row.lesson_type,
        row.link,
        row.room,
        row.groups,
        row.course,
        row.notes,
    ]
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row_index, column)
        _set_cell_value(cell, value)
        _apply_cell_style(cell, style_pack.body_style)
    if style_pack.body_row_height is not None:
        sheet.row_dimensions[row_index].height = style_pack.body_row_height


def _capture_template_styles(template_sheet) -> TemplateStylePack:
    return TemplateStylePack(
        title_style=_snapshot_cell_style(template_sheet["A1"]),
        header_style=_snapshot_cell_style(template_sheet["A2"]),
        body_style=_snapshot_cell_style(template_sheet["A3"]),
        title_row_height=template_sheet.row_dimensions[1].height,
        header_row_height=template_sheet.row_dimensions[2].height,
        body_row_height=template_sheet.row_dimensions[3].height,
    )


def _snapshot_cell_style(cell) -> CellStyleSnapshot:
    return CellStyleSnapshot(
        font=copy(cell.font),
        fill=copy(cell.fill),
        border=copy(cell.border),
        alignment=copy(cell.alignment),
        protection=copy(cell.protection),
        number_format=cell.number_format,
    )


def _apply_cell_style(cell, style: CellStyleSnapshot) -> None:
    cell.font = copy(style.font)
    cell.fill = copy(style.fill)
    cell.border = copy(style.border)
    cell.alignment = copy(style.alignment)
    cell.protection = copy(style.protection)
    cell.number_format = style.number_format


def _set_cell_value(cell, value) -> None:
    cell.value = value
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"


def _export_faculty_label(row: NormalizedRow) -> str:
    return coalesce_label(row.faculty, row.source_name, fallback="unknown faculty")


def _export_program_label(row: NormalizedRow) -> str:
    return coalesce_label(row.program, row.sheet_name, row.source_name, fallback="unknown program")


def _export_sheet_label(row: NormalizedRow) -> str:
    return coalesce_label(row.sheet_name, row.course, row.program, row.source_name, fallback="Аркуш1")


def _sort_key(row: NormalizedRow) -> tuple[int, str, str, str]:
    day_order = {
        "Понеділок": 1,
        "Вівторок": 2,
        "Середа": 3,
        "Четвер": 4,
        "П'ятниця": 5,
        "Субота": 6,
        "Неділя": 7,
    }
    return (day_order.get(row.day, 99), row.start_time, row.week_type, row.subject)


def _truncate_export_name(value: str, *, max_length: int) -> str:
    cleaned = value.strip(" .")
    if not cleaned:
        return "untitled"
    if len(cleaned) <= max_length:
        return cleaned
    trimmed = cleaned[:max_length].rstrip(" ._-")
    return trimmed or "untitled"
