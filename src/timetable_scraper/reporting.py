from __future__ import annotations

from collections import Counter, defaultdict
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from .models import DiscoveryResult, NormalizedRow, SourceConfig, SourceRunSummary
from .utils import ensure_parent, humanize_source_name, json_dumps


TOP_ISSUE_LIMIT = 10
TOP_EXCERPT_LIMIT = 3


def build_source_summaries(
    sources: list[SourceConfig],
    discovery: DiscoveryResult,
    accepted_rows: list[NormalizedRow],
    review_rows: list[NormalizedRow],
    *,
    attempted_assets: Counter[str],
    runtime_issues: dict[str, list[str]],
) -> list[SourceRunSummary]:
    discovered_assets = Counter(asset.source_name for asset in discovery.assets)
    discovery_issues: dict[str, list[str]] = defaultdict(list)
    for issue in discovery.issues:
        reason = issue.reason if not issue.locator else f"{issue.reason} [{issue.locator}]"
        discovery_issues[issue.source_name].append(reason)

    accepted_counts = Counter(row.source_name for row in accepted_rows)
    review_counts = Counter(row.source_name for row in review_rows)
    autofix_counts = Counter(row.source_name for row in [*accepted_rows, *review_rows] if row.autofix_actions)
    review_warning_counts: dict[str, Counter[str]] = defaultdict(Counter)
    review_qa_flag_counts: dict[str, Counter[str]] = defaultdict(Counter)
    for row in review_rows:
        review_warning_counts[row.source_name].update(row.warnings)
        review_qa_flag_counts[row.source_name].update(row.qa_flags)

    summaries: list[SourceRunSummary] = []
    for source in sources:
        source_root_url = source.url or (str(source.path.resolve()) if source.path is not None else "")
        top_review_warnings = [warning for warning, _ in review_warning_counts[source.name].most_common(5)]
        top_review_qa_flags = [flag for flag, _ in review_qa_flag_counts[source.name].most_common(5)]
        status, note = _classify_source(
            accepted_rows=accepted_counts[source.name],
            review_rows=review_counts[source.name],
            discovered_assets=discovered_assets[source.name],
            discovery_issues=discovery_issues[source.name],
            runtime_issues=runtime_issues.get(source.name, []),
            top_review_warnings=top_review_warnings,
            top_review_qa_flags=top_review_qa_flags,
        )
        summaries.append(
            SourceRunSummary(
                source_name=source.name,
                source_root_url=source_root_url,
                status=status,
                accepted_rows=accepted_counts[source.name],
                review_rows=review_counts[source.name],
                autofix_rows=autofix_counts[source.name],
                discovered_assets=discovered_assets[source.name],
                attempted_assets=attempted_assets[source.name],
                discovery_issues=discovery_issues[source.name],
                runtime_issues=runtime_issues.get(source.name, []),
                top_review_warnings=top_review_warnings,
                top_review_qa_flags=top_review_qa_flags,
                note=note,
            )
        )
    return summaries


def write_source_summaries(summaries: list[SourceRunSummary], *, output_dir: Path) -> tuple[Path, Path]:
    json_path = output_dir / "source_summary.json"
    report_path = output_dir / "source_summary.md"
    ensure_parent(json_path)
    payload = [
        {
            "source_name": summary.source_name,
            "source_root_url": summary.source_root_url,
            "status": summary.status,
            "accepted_rows": summary.accepted_rows,
            "review_rows": summary.review_rows,
            "autofix_rows": summary.autofix_rows,
            "discovered_assets": summary.discovered_assets,
            "attempted_assets": summary.attempted_assets,
            "discovery_issues": summary.discovery_issues,
            "runtime_issues": summary.runtime_issues,
            "top_review_warnings": summary.top_review_warnings,
            "top_review_qa_flags": summary.top_review_qa_flags,
            "note": summary.note,
        }
        for summary in summaries
    ]
    json_path.write_text("[\n" + ",\n".join(json_dumps(item) for item in payload) + "\n]\n", encoding="utf-8")
    report_path.write_text(_render_summary_markdown(summaries), encoding="utf-8")
    return json_path, report_path


def write_review_summary(review_rows: list[NormalizedRow], *, output_dir: Path) -> tuple[Path, Path]:
    payload = _build_review_summary_payload(review_rows)
    json_path = output_dir / "review_summary.json"
    xlsx_path = output_dir / "review_summary.xlsx"
    ensure_parent(json_path)
    json_path.write_text(json_dumps(payload) + "\n", encoding="utf-8")
    _write_review_summary_xlsx(payload, xlsx_path)
    return json_path, xlsx_path


def load_previous_source_summaries(output_dir: Path) -> dict[str, dict[str, Any]]:
    path = output_dir / "source_summary.json"
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    if not isinstance(payload, list):
        return {}
    previous: dict[str, dict[str, Any]] = {}
    for item in payload:
        if isinstance(item, dict) and isinstance(item.get("source_name"), str):
            previous[item["source_name"]] = item
    return previous


def write_run_delta(
    summaries: list[SourceRunSummary],
    previous_summaries: dict[str, dict[str, Any]],
    *,
    output_dir: Path,
) -> Path:
    path = output_dir / "run_delta.json"
    current_by_source = {summary.source_name: summary for summary in summaries}
    all_source_names = sorted(set(previous_summaries) | set(current_by_source))

    current_totals = {
        "accepted_rows": sum(summary.accepted_rows for summary in summaries),
        "review_rows": sum(summary.review_rows for summary in summaries),
        "autofix_rows": sum(summary.autofix_rows for summary in summaries),
    }
    previous_totals = {
        "accepted_rows": sum(int(previous_summaries.get(name, {}).get("accepted_rows", 0)) for name in all_source_names),
        "review_rows": sum(int(previous_summaries.get(name, {}).get("review_rows", 0)) for name in all_source_names),
        "autofix_rows": sum(int(previous_summaries.get(name, {}).get("autofix_rows", 0)) for name in all_source_names),
    }

    payload = {
        "previous_run_found": bool(previous_summaries),
        "totals": _delta_block(current_totals, previous_totals),
        "sources": [
            _build_run_delta_source_item(name, current_by_source.get(name), previous_summaries.get(name))
            for name in all_source_names
        ],
    }
    ensure_parent(path)
    path.write_text(json_dumps(payload) + "\n", encoding="utf-8")
    return path


def _build_run_delta_source_item(
    source_name: str,
    current: SourceRunSummary | None,
    previous: dict[str, Any] | None,
) -> dict[str, Any]:
    current_values = {
        "accepted_rows": current.accepted_rows if current is not None else 0,
        "review_rows": current.review_rows if current is not None else 0,
        "autofix_rows": current.autofix_rows if current is not None else 0,
    }
    previous_values = {
        "accepted_rows": int((previous or {}).get("accepted_rows", 0)),
        "review_rows": int((previous or {}).get("review_rows", 0)),
        "autofix_rows": int((previous or {}).get("autofix_rows", 0)),
    }
    return {
        "source_name": source_name,
        "source_root_url": current.source_root_url if current is not None else str((previous or {}).get("source_root_url", "")),
        "status_current": current.status if current is not None else "missing",
        "status_previous": str((previous or {}).get("status", "missing")),
        "delta": _delta_block(current_values, previous_values),
    }


def _delta_block(current_values: dict[str, int], previous_values: dict[str, int]) -> dict[str, dict[str, int]]:
    return {
        key: {
            "current": int(current_values.get(key, 0)),
            "previous": int(previous_values.get(key, 0)),
            "delta": int(current_values.get(key, 0)) - int(previous_values.get(key, 0)),
        }
        for key in ("accepted_rows", "review_rows", "autofix_rows")
    }


def _build_review_summary_payload(review_rows: list[NormalizedRow]) -> dict[str, Any]:
    grouped: dict[str, list[NormalizedRow]] = defaultdict(list)
    for row in review_rows:
        grouped[row.source_name].append(row)

    sources: list[dict[str, Any]] = []
    for source_name, rows in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0])):
        warnings = Counter(warning for row in rows for warning in row.warnings)
        qa_flags = Counter(flag for row in rows for flag in row.qa_flags)
        sources.append(
            {
                "source_name": source_name,
                "source_root_url": next((row.source_root_url for row in rows if row.source_root_url), ""),
                "review_rows": len(rows),
                "warning_counts": dict(warnings.most_common()),
                "qa_flag_counts": dict(qa_flags.most_common()),
                "issues": _build_review_issue_payload(rows, warnings, qa_flags),
            }
        )
    return {
        "total_review_rows": len(review_rows),
        "source_count": len(sources),
        "sources": sources,
    }


def _build_review_issue_payload(
    rows: list[NormalizedRow],
    warnings: Counter[str],
    qa_flags: Counter[str],
) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for issue_name, count in warnings.most_common(TOP_ISSUE_LIMIT):
        issue_rows = [row for row in rows if issue_name in row.warnings]
        issues.append(
            {
                "issue_type": "warning",
                "issue_name": issue_name,
                "count": count,
                "top_raw_excerpts": _top_raw_excerpts(issue_rows),
            }
        )
    for issue_name, count in qa_flags.most_common(TOP_ISSUE_LIMIT):
        issue_rows = [row for row in rows if issue_name in row.qa_flags]
        issues.append(
            {
                "issue_type": "qa_flag",
                "issue_name": issue_name,
                "count": count,
                "top_raw_excerpts": _top_raw_excerpts(issue_rows),
            }
        )
    issues.sort(key=lambda item: (-int(item["count"]), str(item["issue_type"]), str(item["issue_name"])))
    return issues[: TOP_ISSUE_LIMIT * 2]


def _top_raw_excerpts(rows: list[NormalizedRow]) -> list[dict[str, Any]]:
    excerpts = Counter(_normalize_excerpt(row) for row in rows if _normalize_excerpt(row))
    return [
        {"excerpt": excerpt, "count": count}
        for excerpt, count in excerpts.most_common(TOP_EXCERPT_LIMIT)
    ]


def _normalize_excerpt(row: NormalizedRow) -> str:
    text = (row.raw_excerpt or row.subject or row.notes).strip()
    if not text:
        return ""
    return text[:240]


def _write_review_summary_xlsx(payload: dict[str, Any], path: Path) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_headers = ["source_name", "source_root_url", "review_rows", "top_warning", "top_qa_flag"]
    for column, title in enumerate(summary_headers, start=1):
        summary_sheet.cell(1, column).value = title

    issues_sheet = workbook.create_sheet("issues")
    issues_headers = ["source_name", "issue_type", "issue_name", "count", "top_excerpt_1", "top_excerpt_2", "top_excerpt_3"]
    for column, title in enumerate(issues_headers, start=1):
        issues_sheet.cell(1, column).value = title

    row_index = 2
    issue_row_index = 2
    for source in payload["sources"]:
        warning_counts = source["warning_counts"]
        qa_flag_counts = source["qa_flag_counts"]
        top_warning = next(iter(warning_counts), "")
        top_qa_flag = next(iter(qa_flag_counts), "")
        summary_sheet.cell(row_index, 1).value = source["source_name"]
        summary_sheet.cell(row_index, 2).value = source["source_root_url"]
        summary_sheet.cell(row_index, 3).value = source["review_rows"]
        summary_sheet.cell(row_index, 4).value = top_warning
        summary_sheet.cell(row_index, 5).value = top_qa_flag
        row_index += 1

        for issue in source["issues"]:
            excerpts = [item["excerpt"] for item in issue["top_raw_excerpts"]]
            issues_sheet.cell(issue_row_index, 1).value = source["source_name"]
            issues_sheet.cell(issue_row_index, 2).value = issue["issue_type"]
            issues_sheet.cell(issue_row_index, 3).value = issue["issue_name"]
            issues_sheet.cell(issue_row_index, 4).value = issue["count"]
            for offset, excerpt in enumerate(excerpts, start=5):
                issues_sheet.cell(issue_row_index, offset).value = excerpt
            issue_row_index += 1

    ensure_parent(path)
    workbook.save(path)


def _classify_source(
    *,
    accepted_rows: int,
    review_rows: int,
    discovered_assets: int,
    discovery_issues: list[str],
    runtime_issues: list[str],
    top_review_warnings: list[str],
    top_review_qa_flags: list[str],
) -> tuple[str, str]:
    blocker_reason = _find_blocker_reason([*runtime_issues, *discovery_issues])
    if accepted_rows > 0:
        if review_rows > 0:
            return "parsed", f"{review_rows} review rows remain"
        return "parsed", ""
    if blocker_reason:
        return "confirmed-blocker", blocker_reason
    if review_rows > 0:
        incomplete_fields = {"missing_day", "missing_subject", "missing_time"}
        qa_only_blockers = {"garbage_text", "inconsistent_columns", "subject_too_long"}
        if qa_only_blockers & set(top_review_qa_flags) and "garbage_text_subject" in set(top_review_warnings):
            return "confirmed-blocker", "Extracted rows are too noisy to form a reliable timetable"
        if incomplete_fields & set(top_review_warnings):
            return "review-only", "Rows were extracted but key schedule fields are incomplete"
        if qa_only_blockers & set(top_review_qa_flags) and not (incomplete_fields & set(top_review_qa_flags)):
            return "confirmed-blocker", "Extracted rows are too noisy to form a reliable timetable"
        return "review-only", ", ".join(top_review_warnings[:3] or top_review_qa_flags[:3])
    if discovered_assets > 0:
        return "confirmed-blocker", "Assets were discovered but parsers did not yield complete schedule rows"
    return "confirmed-blocker", "No public schedule assets discovered from the official source"


def _find_blocker_reason(reasons: list[str]) -> str:
    for reason in reasons:
        lowered = reason.casefold()
        if "http 500" in lowered or "500 internal server error" in lowered or "500 server error" in lowered:
            return "Official source returns HTTP 500"
        if "http 403" in lowered or "403 client error" in lowered or "blocked" in lowered or "cloudflare" in lowered:
            return "Access to the official source is blocked (HTTP 403 / Cloudflare)"
        if "http 410" in lowered or "410 client error" in lowered or "gone for url" in lowered:
            return "Official schedule asset is no longer publicly available (HTTP 410)"
        if "onedrive public download blocked" in lowered:
            return "OneDrive public download is blocked"
    return ""


def _render_summary_markdown(summaries: list[SourceRunSummary]) -> str:
    lines = ["# Source Summary", ""]
    for status in ("parsed", "review-only", "confirmed-blocker"):
        lines.append(f"## {status}")
        subset = [summary for summary in summaries if summary.status == status]
        if not subset:
            lines.append("- none")
            lines.append("")
            continue
        for summary in subset:
            label = humanize_source_name(summary.source_name) or summary.source_name
            detail = (
                f"accepted={summary.accepted_rows}, review={summary.review_rows}, "
                f"autofix={summary.autofix_rows}, assets={summary.discovered_assets}"
            )
            suffix = f" - {summary.note}" if summary.note else ""
            lines.append(f"- {label} - {detail} - {summary.source_root_url}{suffix}")
        lines.append("")
    return "\n".join(lines).strip() + "\n"
