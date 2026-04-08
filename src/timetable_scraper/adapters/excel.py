from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, cast

import xlrd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..models import FetchedAsset, ParsedDocument, ParsedSheet, RawRecord
from ..normalize import records_from_tabular_rows
from ..utils import (
    DAY_NAMES,
    excerpt_from_values,
    flatten_multiline,
    infer_faculty_from_locator,
    looks_like_roomish_subject_text,
    normalize_day,
    normalize_header,
    parse_time_range,
    parse_time_value,
)

DAY_OR_TIME_HEADERS = {"день", "час"}
FIT_TEACHER_RE = re.compile(
    r"(?i)(?:проф\.?|доц\.?|ас\.?|ст\.викл\.?|викл\.?)|[А-ЯІЇЄҐ][а-яіїєґ'-]+\s+[А-ЯІЇЄҐ]\.\s*[А-ЯІЇЄҐ]\."
)
FIT_ROOM_RE = re.compile(r"(?i)(ауд(?:\.|\b)|корпус|корп(?:\.|\b)|клінік|лаборатор|каб(?:\.|\b)|online|онлайн)")
FIT_LINK_RE = re.compile(
    r"(?i)(https?://|zoom|meet|teams|microsoft|google meet|knu-ua\.zoom|meeting\s*id|passcode|pwd=|id:\s*\d|код(?:[:\s]|\s+доступу))"
)
FIT_WEEK_RE = re.compile(r"(?i)(\[[^\]]+\]|\bч/т\b|\bпо\s+\d{2}\.\d{2}\b|\bз\s+\d{2}\.\d{2}\b|\b(?:i|ii|1|2)\s*тиж)")
FIT_COUNT_RE = re.compile(r"^\d+(?:\.0)?$")
FIT_TRAILING_WEEKS_RE = re.compile(r"\s*(\[[^\]]+\])+\s*$")
FIT_T_COUNT_RE = re.compile(r"\b(\d+)\s*т\b", re.IGNORECASE)
FIT_WEEK_TYPE_PATTERNS = (
    (re.compile(r"(?i)\bч/т\b"), "Через тиждень"),
    (re.compile(r"(?i)\b(?:i|1)\s*тиж"), "Верхній"),
    (re.compile(r"(?i)\b(?:ii|2)\s*тиж"), "Нижній"),
    (re.compile(r"(?i)\bверх"), "Верхній"),
    (re.compile(r"(?i)\bниж"), "Нижній"),
    (re.compile(r"(?i)\bнепар"), "Верхній"),
    (re.compile(r"(?i)\bпарн"), "Нижній"),
)
FIT_LESSON_TYPES = {
    "л": "лекція",
    "лек": "лекція",
    "лекція": "лекція",
    "лаб": "лабораторна",
    "лабораторна": "лабораторна",
    "пр": "практика",
    "практ": "практика",
    "практика": "практика",
    "сем": "семінар",
    "семінар": "семінар",
}

GENERIC_COURSE_RE = re.compile(r"(?i)\b(\d+)\s*курс\b")
GENERIC_GROUP_RE = re.compile(r"(?i)\b(?:група|підгрупа|subgroup)\b")
GENERIC_PARTIAL_TIME_RE = re.compile(r"(?P<start>\d{1,2}[:.]\d{2})\s*[-–—]?\s*$")
GENERIC_END_TIME_RE = re.compile(r"^(?P<end>\d{1,2}[:.]\d{2})$")
GENERIC_HEADER_LABEL_RE = re.compile(
    r"(?iu)^(?:[ivxіvх]+\s*курс|\d+\s*курс|\d+\s*магістр(?:и|ів)?|івт|астрономія|оптика|фізичне матеріало-\s*знавство)$"
)


@dataclass(slots=True)
class GridCell:
    text: str
    kind: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int


def parse_excel_asset(fetched_asset: FetchedAsset) -> ParsedDocument:
    suffix = Path(fetched_asset.resolved_locator.split("::")[-1]).suffix.lower()
    if suffix == ".csv":
        return _parse_csv_asset(fetched_asset)
    if suffix == ".xls":
        return _parse_xls_asset(fetched_asset)
    return _parse_xlsx_asset(fetched_asset)


def _parse_xlsx_asset(fetched_asset: FetchedAsset) -> ParsedDocument:
    workbook = load_workbook(BytesIO(fetched_asset.content), data_only=True)
    faculty = _resolve_faculty(fetched_asset)
    sheets: list[ParsedSheet] = []
    warnings: list[str] = []
    for worksheet in workbook.worksheets:
        if _looks_like_fit_grid_schedule(worksheet):
            records, row_warnings = _parse_fit_grid_schedule_sheet(worksheet, faculty=faculty)
            if records or row_warnings:
                sheets.append(
                    ParsedSheet(
                        sheet_name=worksheet.title,
                        program=flatten_multiline(worksheet.title),
                        faculty=faculty,
                        records=records,
                        warnings=row_warnings,
                    )
                )
                warnings.extend(row_warnings)
            continue
        rows = [
            list(row)
            for row in worksheet.iter_rows(values_only=True)
            if any(cell not in ("", None) for cell in row)
        ]
        if not rows:
            warnings.append(f"Skipped empty sheet '{worksheet.title}'.")
            continue
        program = _extract_program_title(rows, fallback=worksheet.title)
        if _looks_like_generic_grid_schedule(worksheet):
            records, row_warnings = _parse_generic_grid_schedule_sheet(worksheet, faculty=faculty, program=program)
            if not records:
                records, row_warnings = records_from_tabular_rows(
                    rows,
                    program=program,
                    faculty=faculty,
                    sheet_name=worksheet.title,
                )
        else:
            records, row_warnings = records_from_tabular_rows(
                rows,
                program=program,
                faculty=faculty,
                sheet_name=worksheet.title,
            )
        sheets.append(
            ParsedSheet(
                sheet_name=worksheet.title,
                program=program,
                faculty=faculty,
                records=records,
                warnings=row_warnings,
            )
        )
        warnings.extend(row_warnings)
    return ParsedDocument(asset=fetched_asset, sheets=sheets, warnings=warnings)


def _parse_xls_asset(fetched_asset: FetchedAsset) -> ParsedDocument:
    workbook = xlrd.open_workbook(file_contents=fetched_asset.content)
    faculty = _resolve_faculty(fetched_asset)
    sheets: list[ParsedSheet] = []
    warnings: list[str] = []
    for worksheet in workbook.sheets():
        rows = [
            worksheet.row_values(index)
            for index in range(worksheet.nrows)
            if any(value not in ("", None) for value in worksheet.row_values(index))
        ]
        if not rows:
            warnings.append(f"Skipped empty sheet '{worksheet.name}'.")
            continue
        program = _extract_program_title(rows, fallback=worksheet.name)
        records, row_warnings = records_from_tabular_rows(rows, program=program, faculty=faculty, sheet_name=worksheet.name)
        sheets.append(
            ParsedSheet(
                sheet_name=worksheet.name,
                program=program,
                faculty=faculty,
                records=records,
                warnings=row_warnings,
            )
        )
        warnings.extend(row_warnings)
    return ParsedDocument(asset=fetched_asset, sheets=sheets, warnings=warnings)


def _parse_csv_asset(fetched_asset: FetchedAsset) -> ParsedDocument:
    reader = csv.reader(StringIO(fetched_asset.content.decode("utf-8-sig", "ignore")))
    rows = [row for row in reader if any(cell not in ("", None) for cell in row)]
    faculty = _resolve_faculty(fetched_asset)
    program = _extract_program_title(rows, fallback=fetched_asset.asset.display_name)
    records, warnings = records_from_tabular_rows(rows, program=program, faculty=faculty, sheet_name="Аркуш1")
    return ParsedDocument(
        asset=fetched_asset,
        sheets=[ParsedSheet(sheet_name="Аркуш1", program=program, faculty=faculty, records=records, warnings=warnings)],
        warnings=warnings,
    )


def _extract_program_title(rows: list[list[Any]], fallback: str) -> str:
    for row in rows[:3]:
        values = [flatten_multiline(cell) for cell in row if flatten_multiline(cell)]
        if len(values) == 1:
            return values[0]
    return flatten_multiline(fallback)


def _resolve_faculty(fetched_asset: FetchedAsset) -> str:
    locator = fetched_asset.asset.source_root_url or fetched_asset.asset.locator
    source_name = fetched_asset.asset.source_name.casefold()
    if "fit.knu.ua" in locator or "fit" in source_name or "фіт" in source_name:
        return "Факультет інформаційних технологій"
    faculty = infer_faculty_from_locator(locator)
    if faculty == "docs.google.com":
        return "Невідомий факультет"
    return faculty


def _looks_like_fit_grid_schedule(worksheet: Worksheet) -> bool:
    if worksheet.max_column < 4 or worksheet.max_row < 10:
        return False
    header_row = [normalize_header(worksheet.cell(2, column).value) for column in range(1, min(worksheet.max_column, 16) + 1)]
    return "день" in header_row and "час" in header_row


def _parse_fit_grid_schedule_sheet(worksheet: Worksheet, *, faculty: str) -> tuple[list[RawRecord], list[str]]:
    merged_lookup = _build_merged_lookup(worksheet)
    header_context = _build_fit_header_context(worksheet, merged_lookup)
    schedule_columns = [column for column, header in header_context["course"].items() if header]
    slot_starts = [row for row in range(5, worksheet.max_row + 1) if parse_time_range(_expanded_value(worksheet, row, 2, merged_lookup))[0]]
    if not slot_starts:
        return [], [f"Could not detect FIT-style schedule slots in sheet '{worksheet.title}'."]
    records: list[RawRecord] = []
    warnings: list[str] = []
    current_day = ""
    for slot_start in slot_starts:
        day = normalize_day(_expanded_value(worksheet, slot_start, 1, merged_lookup)) or current_day
        if day:
            current_day = day
        time_text = _expanded_value(worksheet, slot_start, 2, merged_lookup)
        start_time, end_time = parse_time_range(time_text)
        if not start_time or not end_time:
            continue
        block_cells = _collect_fit_block_cells(worksheet, slot_start, merged_lookup, schedule_columns)
        subject_cells = [cell for cell in block_cells if cell.kind == "subject"]
        for subject_cell in subject_cells:
            groups = _compose_fit_groups(header_context, subject_cell.min_col, subject_cell.max_col)
            course = _compose_fit_courses(header_context, subject_cell.min_col, subject_cell.max_col)
            teacher, room, link, notes = _collect_fit_metadata(subject_cell, block_cells)
            subject, lesson_type, subject_notes = _split_fit_subject(subject_cell.text)
            merged_notes = _merge_notes(subject_notes, notes)
            week_type = _infer_fit_week_type(subject_cell.text, subject_notes, notes)
            values = {
                "program": flatten_multiline(worksheet.title),
                "faculty": faculty,
                "week_type": week_type,
                "day": day,
                "start_time": start_time,
                "end_time": end_time,
                "subject": subject,
                "teacher": teacher,
                "lesson_type": lesson_type,
                "link": link,
                "room": room,
                "groups": groups,
                "course": course,
                "notes": merged_notes,
            }
            if not values["subject"]:
                continue
            records.append(
                RawRecord(
                    values=values,
                    row_index=slot_start,
                    sheet_name=worksheet.title,
                    raw_excerpt=excerpt_from_values(values),
                )
            )
    if not records:
        warnings.append(f"Could not extract FIT-style rows from sheet '{worksheet.title}'.")
    return records, warnings


def _build_fit_header_context(worksheet: Worksheet, merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]]) -> dict[str, dict[int, str]]:
    course_map: dict[int, str] = {}
    group_map: dict[int, str] = {}
    subgroup_map: dict[int, str] = {}
    current_course = ""
    current_group = ""
    current_subgroup = ""
    for column in range(3, worksheet.max_column + 1):
        raw_course = _expanded_value(worksheet, 2, column, merged_lookup)
        normalized_course = normalize_header(raw_course)
        if normalized_course in DAY_OR_TIME_HEADERS:
            current_course = ""
            current_group = ""
            current_subgroup = ""
        elif raw_course:
            current_course = raw_course
        raw_group = _expanded_value(worksheet, 3, column, merged_lookup)
        if raw_group:
            current_group = raw_group
        raw_subgroup = _expanded_value(worksheet, 4, column, merged_lookup)
        if raw_subgroup:
            current_subgroup = raw_subgroup
        course_map[column] = current_course
        group_map[column] = current_group
        subgroup_map[column] = current_subgroup
    return {"course": course_map, "group": group_map, "subgroup": subgroup_map}


def _collect_fit_block_cells(
    worksheet: Worksheet,
    slot_start: int,
    merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
    schedule_columns: list[int],
) -> list[GridCell]:
    block_end = min(slot_start + 7, worksheet.max_row)
    seen_anchors: set[tuple[int, int, int, int]] = set()
    cells: list[GridCell] = []
    for row in range(slot_start, block_end + 1):
        for column in schedule_columns:
            value = worksheet.cell(row, column).value
            if value in ("", None):
                continue
            anchor = merged_lookup.get((row, column), (row, column, row, column))
            if anchor in seen_anchors:
                continue
            seen_anchors.add(anchor)
            text = flatten_multiline(worksheet.cell(anchor[0], anchor[1]).value)
            kind = _classify_fit_cell(text)
            if kind == "ignore":
                continue
            cells.append(
                GridCell(
                    text=text,
                    kind=kind,
                    min_row=anchor[0],
                    max_row=anchor[2],
                    min_col=anchor[1],
                    max_col=anchor[3],
                )
            )
    return cells


def _classify_fit_cell(text: str) -> str:
    cleaned = flatten_multiline(text)
    if not cleaned or FIT_COUNT_RE.fullmatch(cleaned):
        return "ignore"
    normalized = normalize_header(cleaned)
    if normalized in DAY_OR_TIME_HEADERS or normalize_day(cleaned) in DAY_NAMES.values():
        return "ignore"
    if parse_time_range(cleaned)[0]:
        return "ignore"
    if FIT_LINK_RE.search(cleaned):
        return "link"
    if looks_like_roomish_subject_text(cleaned):
        return "room"
    if FIT_ROOM_RE.search(cleaned):
        return "room"
    if FIT_TEACHER_RE.search(cleaned):
        return "teacher"
    if FIT_WEEK_RE.search(cleaned) and not _looks_like_subject_text(cleaned):
        return "week"
    if _looks_like_fit_noise_text(cleaned):
        return "ignore"
    if not any(character.isalpha() for character in cleaned):
        return "ignore"
    return "subject"


def _looks_like_subject_text(text: str) -> bool:
    lowered = text.casefold()
    return any(marker in lowered for marker in ("(", "лаб", "лек", "пр", "сем", "основи", "архітект", "матем", "комп", "кібер", "інозем"))


def _looks_like_fit_noise_text(text: str) -> bool:
    cleaned = flatten_multiline(text)
    compact = cleaned.replace(" ", "")
    if not cleaned:
        return True
    if len(cleaned) <= 3 and cleaned.endswith("."):
        return True
    if normalize_header(cleaned) in {"ст", "ас", "доц", "проф", "викл"}:
        return True
    if compact.startswith(".") and "microsoft" in compact.casefold():
        return True
    if re.fullmatch(r"[A-Za-z0-9._=+-]{6,}", compact):
        has_mixed_code = any(character.isdigit() for character in compact) and any(character.isalpha() for character in compact)
        if has_mixed_code or compact.casefold().startswith(("pwd", "id")):
            return True
    return False


def _collect_fit_metadata(subject_cell: GridCell, block_cells: list[GridCell]) -> tuple[str, str, str, list[str]]:
    teacher_parts: list[str] = []
    room_parts: list[str] = []
    link_parts: list[str] = []
    note_parts: list[str] = []
    for cell in block_cells:
        if cell is subject_cell or cell.kind == "subject":
            continue
        if not _column_ranges_overlap(subject_cell, cell):
            continue
        cleaned = flatten_multiline(cell.text)
        if cell.kind == "teacher":
            teacher_text, extracted_notes = _split_bracket_notes(cleaned)
            if teacher_text:
                teacher_parts.append(teacher_text)
            note_parts.extend(extracted_notes)
        elif cell.kind == "room":
            room_parts.append(cleaned)
        elif cell.kind == "link":
            link_parts.append(cleaned)
        elif cell.kind == "week":
            note_parts.append(cleaned)
    return (
        _join_unique(teacher_parts),
        _join_unique(room_parts),
        _join_unique(link_parts, separator=" "),
        _unique_list(note_parts),
    )


def _split_fit_subject(text: str) -> tuple[str, str, list[str]]:
    cleaned = flatten_multiline(text)
    notes: list[str] = []
    trailing_weeks = FIT_TRAILING_WEEKS_RE.findall(cleaned)
    if trailing_weeks:
        notes.extend(trailing_weeks)
        cleaned = FIT_TRAILING_WEEKS_RE.sub("", cleaned).strip()
    lesson_type = ""
    lesson_type_match = re.search(r"\(([^()]{1,12})\)", cleaned)
    if lesson_type_match:
        raw_lesson_type = lesson_type_match.group(1).strip().casefold().rstrip(".")
        lesson_type = FIT_LESSON_TYPES.get(raw_lesson_type, lesson_type_match.group(1).strip())
        cleaned = cleaned.replace(lesson_type_match.group(0), "").strip()
    t_count_match = FIT_T_COUNT_RE.search(cleaned)
    if t_count_match:
        notes.append(f"T={t_count_match.group(1)}")
        cleaned = FIT_T_COUNT_RE.sub("", cleaned).strip()
    if lesson_type:
        cleaned = re.sub(r"\s{2,}", " ", cleaned).strip(" -/;")
    if _looks_like_fit_noise_text(cleaned) or looks_like_roomish_subject_text(cleaned):
        return "", lesson_type, _unique_list(notes)
    return cleaned or flatten_multiline(text), lesson_type, _unique_list(notes)


def _split_bracket_notes(text: str) -> tuple[str, list[str]]:
    notes = re.findall(r"\[[^\]]+\]", text)
    cleaned = re.sub(r"\[[^\]]+\]", "", text).strip()
    return cleaned, _unique_list(notes)


def _infer_fit_week_type(subject_text: str, subject_notes: list[str], metadata_notes: list[str]) -> str:
    candidates = [subject_text, *subject_notes, *metadata_notes]
    for candidate in candidates:
        cleaned = flatten_multiline(candidate)
        if not cleaned:
            continue
        for pattern, week_type in FIT_WEEK_TYPE_PATTERNS:
            if pattern.search(cleaned):
                return week_type
    return ""


def _compose_fit_groups(header_context: dict[str, dict[int, str]], min_col: int, max_col: int) -> str:
    subgroup_values = _unique_list(
        header_context["subgroup"].get(column, "")
        for column in range(min_col, max_col + 1)
        if header_context["subgroup"].get(column, "")
    )
    if subgroup_values:
        return "; ".join(subgroup_values)
    group_values = _unique_list(
        header_context["group"].get(column, "")
        for column in range(min_col, max_col + 1)
        if header_context["group"].get(column, "")
    )
    return "; ".join(group_values)


def _compose_fit_courses(header_context: dict[str, dict[int, str]], min_col: int, max_col: int) -> str:
    course_values: list[str] = []
    for column in range(min_col, max_col + 1):
        header = header_context["course"].get(column, "")
        match = re.search(r"\b(\d+)\s*курс\b", header.casefold())
        if match:
            course_values.append(match.group(1))
    return "; ".join(_unique_list(course_values))


def _merge_notes(*note_groups: list[str]) -> str:
    notes: list[str] = []
    for group in note_groups:
        notes.extend(group)
    return "; ".join(_unique_list(notes))


def _join_unique(values: list[str], *, separator: str = "; ") -> str:
    return separator.join(_unique_list(values))


def _unique_list(values: Any) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        cleaned = flatten_multiline(value)
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        result.append(cleaned)
    return result


def _expanded_value(
    worksheet: Worksheet,
    row: int,
    column: int,
    merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
) -> str:
    direct = flatten_multiline(worksheet.cell(row, column).value)
    if direct:
        return direct
    anchor = merged_lookup.get((row, column))
    if anchor is None:
        return ""
    return flatten_multiline(worksheet.cell(anchor[0], anchor[1]).value)


def _build_merged_lookup(worksheet: Worksheet) -> dict[tuple[int, int], tuple[int, int, int, int]]:
    lookup: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    for merged_range in worksheet.merged_cells.ranges:
        anchor = (merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                lookup[(row, column)] = anchor
    return lookup


def _column_ranges_overlap(left: GridCell, right: GridCell) -> bool:
    return not (left.max_col < right.min_col or right.max_col < left.min_col)


def _looks_like_generic_grid_schedule(worksheet: Worksheet) -> bool:
    if worksheet.max_column < 3 or worksheet.max_row < 8:
        return False
    merged_lookup = _build_merged_lookup(worksheet)
    day_hits = 0
    time_hits = 0
    for row in range(1, min(worksheet.max_row, 150) + 1):
        for column in range(1, min(worksheet.max_column, 4) + 1):
            text = _expanded_value(worksheet, row, column, merged_lookup)
            if not text:
                continue
            if _normalize_generic_day(text) in DAY_NAMES.values():
                day_hits += 1
            if _parse_generic_time_text(text)[0] or parse_time_value(text) or GENERIC_PARTIAL_TIME_RE.search(text) or GENERIC_END_TIME_RE.search(text):
                time_hits += 1
    return day_hits >= 2 and time_hits >= 2


def _parse_generic_grid_schedule_sheet(
    worksheet: Worksheet,
    *,
    faculty: str,
    program: str,
) -> tuple[list[RawRecord], list[str]]:
    merged_lookup = _build_merged_lookup(worksheet)
    day_col, time_col, first_day_row = _detect_generic_axis_columns(worksheet, merged_lookup)
    if not day_col or not time_col or not first_day_row:
        return [], [f"Could not detect grid schedule axes in sheet '{worksheet.title}'."]
    schedule_min_col = max(day_col, time_col) + 1
    cells = _collect_generic_grid_cells(
        worksheet,
        merged_lookup,
        first_day_row=first_day_row,
        day_col=day_col,
        time_col=time_col,
        schedule_min_col=schedule_min_col,
    )
    day_cells = [cell for cell in cells if cell.kind == "day"]
    time_cells = _build_generic_time_cells(worksheet, merged_lookup, time_col=time_col, first_day_row=first_day_row)
    subject_cells = [cell for cell in cells if cell.kind == "subject" and cell.min_col >= schedule_min_col]
    header_context = _build_generic_header_context(
        worksheet,
        merged_lookup,
        first_day_row=first_day_row,
        schedule_min_col=schedule_min_col,
    )
    records: list[RawRecord] = []
    warnings: list[str] = []
    seen_keys: set[tuple[str, str, str, str, str, str]] = set()
    for subject_cell in subject_cells:
        day = _find_generic_day(subject_cell, day_cells, schedule_min_col=schedule_min_col)
        start_time, end_time = _find_generic_time_range(subject_cell, time_cells)
        if not day or not start_time or not end_time:
            continue
        subject, lesson_type, subject_notes = _split_fit_subject(subject_cell.text)
        subject = flatten_multiline(subject)
        if not subject or len(subject) < 3 or not _is_probable_generic_subject(subject):
            continue
        teacher, room, link, notes = _collect_generic_metadata(subject_cell, cells)
        groups = _compose_generic_groups(header_context, subject_cell.min_col, subject_cell.max_col)
        course = _compose_generic_courses(header_context, subject_cell.min_col, subject_cell.max_col)
        values = {
            "program": program,
            "faculty": faculty,
            "week_type": _infer_fit_week_type(subject_cell.text, subject_notes, notes),
            "day": day,
            "start_time": start_time,
            "end_time": end_time,
            "subject": subject,
            "teacher": teacher,
            "lesson_type": lesson_type,
            "link": link,
            "room": room,
            "groups": groups,
            "course": course,
            "notes": _merge_notes(subject_notes, notes),
        }
        key = (
            values["day"],
            values["start_time"],
            values["end_time"],
            values["subject"],
            values["groups"],
            values["course"],
        )
        if key in seen_keys:
            continue
        seen_keys.add(key)
        records.append(
            RawRecord(
                values=values,
                row_index=subject_cell.min_row,
                sheet_name=worksheet.title,
                raw_excerpt=excerpt_from_values(values),
            )
        )
    if not records:
        warnings.append(f"Could not extract generic grid rows from sheet '{worksheet.title}'.")
    return records, warnings


def _detect_generic_axis_columns(
    worksheet: Worksheet,
    merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
) -> tuple[int, int, int]:
    day_scores: dict[int, int] = {}
    time_scores: dict[int, int] = {}
    first_day_row = 0
    for row in range(1, min(worksheet.max_row, 150) + 1):
        for column in range(1, min(worksheet.max_column, 4) + 1):
            text = _expanded_value(worksheet, row, column, merged_lookup)
            if not text:
                continue
            if _normalize_generic_day(text) in DAY_NAMES.values():
                day_scores[column] = day_scores.get(column, 0) + 1
                if first_day_row == 0 or row < first_day_row:
                    first_day_row = row
            if _parse_generic_time_text(text)[0] or parse_time_value(text) or GENERIC_PARTIAL_TIME_RE.search(text) or GENERIC_END_TIME_RE.search(text):
                time_scores[column] = time_scores.get(column, 0) + 1
    if not day_scores or not time_scores:
        return 0, 0, 0
    day_col = min(day_scores, key=lambda column: (-day_scores[column], column))
    time_col = min(time_scores, key=lambda column: (-time_scores[column], column))
    return day_col, time_col, first_day_row


def _collect_generic_grid_cells(
    worksheet: Worksheet,
    merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
    *,
    first_day_row: int,
    day_col: int,
    time_col: int,
    schedule_min_col: int,
) -> list[GridCell]:
    seen_anchors: set[tuple[int, int, int, int]] = set()
    cells: list[GridCell] = []
    for row in range(first_day_row, worksheet.max_row + 1):
        for column in range(1, worksheet.max_column + 1):
            anchor = merged_lookup.get((row, column), (row, column, row, column))
            if anchor in seen_anchors:
                continue
            seen_anchors.add(anchor)
            text = flatten_multiline(worksheet.cell(anchor[0], anchor[1]).value)
            kind = _classify_generic_cell(text)
            if kind == "ignore":
                continue
            if kind not in {"day", "time"} and _is_generic_header_row(
                worksheet,
                merged_lookup,
                row=anchor[0],
                day_col=day_col,
                time_col=time_col,
            ):
                continue
            if anchor[1] < schedule_min_col and kind not in {"day", "time"}:
                continue
            cells.append(
                GridCell(
                    text=text,
                    kind=kind,
                    min_row=anchor[0],
                    max_row=anchor[2],
                    min_col=anchor[1],
                    max_col=anchor[3],
                )
            )
    return cells


def _is_generic_header_row(
    worksheet: Worksheet,
    merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
    *,
    row: int,
    day_col: int,
    time_col: int,
) -> bool:
    current_day = _normalize_generic_day(_expanded_value(worksheet, row, day_col, merged_lookup))
    current_time = _parse_generic_time_text(_expanded_value(worksheet, row, time_col, merged_lookup))[0] or parse_time_value(
        _expanded_value(worksheet, row, time_col, merged_lookup)
    )
    if current_day in DAY_NAMES.values() or current_time:
        return False
    next_day = _normalize_generic_day(_expanded_value(worksheet, row + 1, day_col, merged_lookup))
    next_time = _parse_generic_time_text(_expanded_value(worksheet, row + 1, time_col, merged_lookup))[0] or parse_time_value(
        _expanded_value(worksheet, row + 1, time_col, merged_lookup)
    )
    return next_day in DAY_NAMES.values() and bool(next_time)


def _classify_generic_cell(text: str) -> str:
    cleaned = flatten_multiline(text)
    if not cleaned or FIT_COUNT_RE.fullmatch(cleaned):
        return "ignore"
    normalized = normalize_header(cleaned)
    if GENERIC_HEADER_LABEL_RE.fullmatch(normalized):
        return "ignore"
    if _normalize_generic_day(cleaned) in DAY_NAMES.values():
        return "day"
    if _parse_generic_time_text(cleaned)[0] or GENERIC_PARTIAL_TIME_RE.search(cleaned) or GENERIC_END_TIME_RE.search(cleaned):
        return "time"
    if normalized in {"дні", "години", "день", "час", "час занять", "курс", "група", "підгрупа"}:
        return "ignore"
    if FIT_LINK_RE.search(cleaned):
        return "link"
    if looks_like_roomish_subject_text(cleaned):
        return "room"
    if FIT_ROOM_RE.search(cleaned) and len(cleaned) <= 180:
        return "room"
    if FIT_TEACHER_RE.search(cleaned) and len(cleaned) <= 220:
        return "teacher"
    if not any(character.isalpha() for character in cleaned):
        return "ignore"
    return "subject"


def _build_generic_time_cells(
    worksheet: Worksheet,
    merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
    *,
    time_col: int,
    first_day_row: int,
) -> list[GridCell]:
    cells: list[GridCell] = []
    seen_rows: set[int] = set()
    for row in range(first_day_row, worksheet.max_row + 1):
        if row in seen_rows:
            continue
        anchor = merged_lookup.get((row, time_col), (row, time_col, row, time_col))
        text = _expanded_value(worksheet, row, time_col, merged_lookup)
        if not text:
            continue
        start_time, end_time = _parse_generic_time_text(text)
        if start_time and end_time:
            for used_row in range(anchor[0], anchor[2] + 1):
                seen_rows.add(used_row)
            cells.append(
                GridCell(
                    text=f"{start_time}-{end_time}",
                    kind="time",
                    min_row=anchor[0],
                    max_row=anchor[2],
                    min_col=time_col,
                    max_col=time_col,
                )
            )
            continue
        cleaned_text = flatten_multiline(text)
        start_match = GENERIC_PARTIAL_TIME_RE.search(text)
        next_row = anchor[2] + 1
        next_text = _expanded_value(worksheet, next_row, time_col, merged_lookup) if next_row <= worksheet.max_row else ""
        next_match = GENERIC_END_TIME_RE.search(next_text)
        if start_match and any(separator in cleaned_text for separator in ("-", "–", "—")) and next_match:
            start_time = parse_time_value(start_match.group("start"))
            end_time = parse_time_value(next_match.group("end"))
            next_anchor = (
                merged_lookup.get((next_row, time_col), (next_row, time_col, next_row, time_col))
                if next_row <= worksheet.max_row
                else (next_row, time_col, next_row, time_col)
            )
            for used_row in range(anchor[0], max(anchor[2], next_anchor[2]) + 1):
                seen_rows.add(used_row)
            cells.append(
                GridCell(
                    text=f"{start_time}-{end_time}",
                    kind="time",
                    min_row=anchor[0],
                    max_row=max(anchor[2], next_anchor[2]),
                    min_col=time_col,
                    max_col=time_col,
                )
            )
            continue
        single_time = parse_time_value(text)
        next_single_time, next_time_row = _find_next_distinct_time(
            worksheet,
            merged_lookup,
            time_col=time_col,
            start_row=anchor[2] + 1,
            current_time=single_time,
        )
        if single_time and next_single_time and not any(separator in cleaned_text for separator in ("-", "–", "—")):
            max_row = max(anchor[2], next_time_row - 1)
            for used_row in range(anchor[0], max_row + 1):
                seen_rows.add(used_row)
            cells.append(
                GridCell(
                    text=f"{single_time}-{_subtract_minutes(next_single_time, 10)}",
                    kind="time",
                    min_row=anchor[0],
                    max_row=max_row,
                    min_col=time_col,
                    max_col=time_col,
                )
            )
            continue
        if single_time and not any(separator in cleaned_text for separator in ("-", "–", "—")):
            for used_row in range(anchor[0], anchor[2] + 1):
                seen_rows.add(used_row)
            cells.append(
                GridCell(
                    text=f"{single_time}-{_add_minutes(single_time, 80)}",
                    kind="time",
                    min_row=anchor[0],
                    max_row=anchor[2],
                    min_col=time_col,
                    max_col=time_col,
                )
            )
    return cells


def _build_generic_header_context(
    worksheet: Worksheet,
    merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
    *,
    first_day_row: int,
    schedule_min_col: int,
) -> dict[str, dict[int, list[str] | str]]:
    header_map: dict[int, list[str]] = {}
    course_map: dict[int, str] = {}
    group_map: dict[int, str] = {}
    for column in range(schedule_min_col, worksheet.max_column + 1):
        values: list[str] = []
        for row in range(1, first_day_row):
            text = _expanded_value(worksheet, row, column, merged_lookup)
            cleaned = flatten_multiline(text)
            if not cleaned:
                continue
            if normalize_day(cleaned) in DAY_NAMES.values():
                continue
            if _parse_generic_time_text(cleaned)[0]:
                continue
            if cleaned in values:
                continue
            values.append(cleaned)
        header_map[column] = values
        course_values = [match.group(1) for value in values for match in [GENERIC_COURSE_RE.search(normalize_header(value))] if match]
        group_values = [value for value in values if GENERIC_GROUP_RE.search(normalize_header(value))]
        course_map[column] = "; ".join(_unique_list(course_values))
        group_map[column] = "; ".join(_unique_list(group_values))
    return {
        "headers": cast(dict[int, list[str] | str], header_map),
        "course": cast(dict[int, list[str] | str], course_map),
        "group": cast(dict[int, list[str] | str], group_map),
    }


def _find_generic_day(subject_cell: GridCell, day_cells: list[GridCell], *, schedule_min_col: int) -> str:
    if not day_cells:
        return ""
    overlapping_column_days = [
        cell
        for cell in day_cells
        if _column_ranges_overlap(subject_cell, cell) and _row_gap(subject_cell, cell) <= 2
    ]
    if overlapping_column_days:
        chosen = min(overlapping_column_days, key=lambda cell: (_row_gap(subject_cell, cell), abs(subject_cell.min_row - cell.min_row)))
        return _normalize_generic_day(chosen.text)
    axis_days = [cell for cell in day_cells if cell.max_col < schedule_min_col and _row_gap(subject_cell, cell) <= 3]
    if axis_days:
        chosen = min(axis_days, key=lambda cell: (_row_gap(subject_cell, cell), abs(subject_cell.min_row - cell.min_row)))
        return _normalize_generic_day(chosen.text)
    chosen = min(day_cells, key=lambda cell: (_row_gap(subject_cell, cell), abs(subject_cell.min_row - cell.min_row)))
    return _normalize_generic_day(chosen.text)


def _find_generic_time_range(subject_cell: GridCell, time_cells: list[GridCell]) -> tuple[str, str]:
    if not time_cells:
        return "", ""
    relevant = [cell for cell in time_cells if _row_gap(subject_cell, cell) <= 2]
    if not relevant:
        return "", ""
    min_gap = min(_row_gap(subject_cell, cell) for cell in relevant)
    threshold = 0 if min_gap == 0 else 1 if min_gap == 1 else 2
    cluster = [cell for cell in relevant if _row_gap(subject_cell, cell) <= threshold]
    ranges = [parse_time_range(cell.text) for cell in cluster]
    start_values = [start for start, _ in ranges if start]
    end_values = [end for _, end in ranges if end]
    if not start_values or not end_values:
        return "", ""
    return min(start_values), max(end_values)


def _collect_generic_metadata(subject_cell: GridCell, cells: list[GridCell]) -> tuple[str, str, str, list[str]]:
    teacher_parts: list[str] = []
    room_parts: list[str] = []
    link_parts: list[str] = []
    note_parts: list[str] = []
    for cell in cells:
        if cell is subject_cell or cell.kind in {"day", "time", "subject"}:
            continue
        if not _column_ranges_overlap(subject_cell, cell):
            continue
        if _row_gap(subject_cell, cell) > 2:
            continue
        cleaned = flatten_multiline(cell.text)
        if cell.kind == "teacher":
            teacher_parts.append(cleaned)
        elif cell.kind == "room":
            room_parts.append(cleaned)
        elif cell.kind == "link":
            link_parts.append(cleaned)
        else:
            note_parts.append(cleaned)
    return (
        _join_unique(teacher_parts),
        _join_unique(room_parts),
        _join_unique(link_parts, separator=" "),
        _unique_list(note_parts),
    )


def _compose_generic_groups(header_context: dict[str, dict[int, list[str] | str]], min_col: int, max_col: int) -> str:
    group_values: list[str] = []
    fallback_headers: list[str] = []
    for column in range(min_col, max_col + 1):
        group_value = flatten_multiline(header_context["group"].get(column, ""))
        if group_value:
            group_values.append(group_value)
        for header in header_context["headers"].get(column, []):
            normalized = normalize_header(header)
            if GENERIC_COURSE_RE.search(normalized):
                continue
            if len(header) > 80:
                continue
            fallback_headers.append(header)
    values = group_values or fallback_headers
    return "; ".join(_unique_list(values))


def _compose_generic_courses(header_context: dict[str, dict[int, list[str] | str]], min_col: int, max_col: int) -> str:
    course_values: list[str] = []
    for column in range(min_col, max_col + 1):
        course_value = flatten_multiline(header_context["course"].get(column, ""))
        if course_value:
            course_values.append(course_value)
    return "; ".join(_unique_list(course_values))


def _parse_generic_time_text(text: str) -> tuple[str, str]:
    start_time, end_time = parse_time_range(text)
    if start_time and end_time:
        return start_time, end_time
    cleaned = flatten_multiline(text)
    match = GENERIC_PARTIAL_TIME_RE.search(cleaned)
    if match and "-" in cleaned:
        return parse_time_value(match.group("start")), ""
    return "", ""


def _row_overlap(left: GridCell, right: GridCell) -> int:
    return max(0, min(left.max_row, right.max_row) - max(left.min_row, right.min_row) + 1)


def _row_gap(left: GridCell, right: GridCell) -> int:
    if _row_overlap(left, right):
        return 0
    if left.max_row < right.min_row:
        return right.min_row - left.max_row
    return left.min_row - right.max_row


def _subtract_minutes(time_value: str, minutes: int) -> str:
    hours, mins = [int(part) for part in time_value.split(":", 1)]
    total = hours * 60 + mins - minutes
    total %= 24 * 60
    return f"{total // 60:02d}:{total % 60:02d}"


def _add_minutes(time_value: str, minutes: int) -> str:
    hours, mins = [int(part) for part in time_value.split(":", 1)]
    total = hours * 60 + mins + minutes
    total %= 24 * 60
    return f"{total // 60:02d}:{total % 60:02d}"


def _find_next_distinct_time(
    worksheet: Worksheet,
    merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
    *,
    time_col: int,
    start_row: int,
    current_time: str,
) -> tuple[str, int]:
    for row in range(start_row, worksheet.max_row + 1):
        candidate = parse_time_value(_expanded_value(worksheet, row, time_col, merged_lookup))
        if candidate and candidate != current_time:
            return candidate, row
    return "", worksheet.max_row + 1


def _normalize_generic_day(text: str) -> str:
    cleaned = flatten_multiline(text)
    direct = normalize_day(cleaned)
    if direct in DAY_NAMES.values():
        return direct
    collapsed = cleaned.replace(" ", "")
    return normalize_day(collapsed)


def _is_probable_generic_subject(text: str) -> bool:
    normalized = normalize_header(text)
    if not normalized:
        return False
    if GENERIC_GROUP_RE.search(normalized):
        return False
    if normalized.startswith(("теоретичне навчання", "розклад занять", "графік", "списки груп", "з ", "до ")):
        return False
    if normalize_day(text) in DAY_NAMES.values():
        return False
    return True
