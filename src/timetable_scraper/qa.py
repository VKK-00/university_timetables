from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook

from .models import NormalizedRow, WorkbookQaSheetSummary, WorkbookQaSummary
from .utils import (
    coalesce_program_label,
    coalesce_label,
    count_program_codes,
    contains_link_text,
    humanize_source_name,
    infer_asset_label_from_locator,
    ensure_parent,
    json_dumps,
    looks_like_admin_text,
    looks_like_bad_program_label,
    normalize_program_candidate,
    looks_like_forbidden_subject_text,
    looks_like_technical_label,
    looks_like_garbage_text,
    looks_like_room_text,
    looks_like_roomish_subject_text,
    looks_like_service_text,
    looks_like_teacher_text,
    normalize_service_tokens,
)


BODY_COLUMN_COUNT = 12
HARD_FAIL_FLAGS = {
    "bad_program_label",
    "missing_day",
    "missing_time",
    "missing_subject",
    "subject_contains_room",
    "subject_contains_teacher",
    "subject_contains_link",
    "subject_too_long",
    "teacher_too_long",
    "garbage_text",
    "inconsistent_columns",
    "implausible_time",
    "service_text_subject",
}
LESSON_TEXT_RE = re.compile(r"(?iu)\((?:лек|прак|сем|лаб|lek|sem|prac)")
TRAILING_ROOM_RE = re.compile(r"(?iu)(?:\s*/\s*|\s+)(?:\d{3,4}[А-ЯІЇЄҐA-Z]?|[А-ЯІЇЄҐA-Z]-?\d{2,4}|(?:хімічний|географічний)\s+ф-?т\s+\d{2,4})$")
EXTENDED_DURATION_SUBJECT_RE = re.compile(
    r"(?iu)\b(?:кваліфікаційн\w+\s+робот\w+|захист\w+|атестаці\w+|dissertation|thesis)\b"
)
ABBREVIATED_SUBJECT_RE = re.compile(r"(?iu)^(?:ст|ас|доц|проф|викл)\.?$")
LOWERCASE_DOTTED_VALID_SUBJECT_RE = re.compile(r"(?iu)^[а-яіїєґ]{4,}(?:\.[а-яіїєґ]{2,}){1,}$")


LONG_PRACTICE_SUBJECT_RE = re.compile(
    "(?iu)\\b(?:\\u043f\\u0440\\u0430\\u043a\\u0442\\u0438\\u043a\\w+|internship|practice)\\b"
)
TINY_BAD_PROGRAM_PATTERNS = (
    re.compile(r"(?iu)^розклад\b"),
    re.compile(r"(?iu)^начитка!?$"),
    re.compile(r"(?iu)^постій"),
    re.compile(r"(?iu)^увага"),
    re.compile(r"(?iu)^початок\s+занять"),
    re.compile(r"(?iu)^навчання\s+з\s+використанням"),
    re.compile(r"(?iu)^\d+\s*магістр\w*$"),
    re.compile(r"(?iu)^\d+\s*курс\s*\([^)]{2,}\)$"),
    re.compile(r"(?iu)^.+\s+\((?:л|пр|лаб|сем)\)$"),
)
DROP_REVIEW_SERVICE_RE = re.compile(
    r"(?iu)^(?:день\s+самост[іi]йної\s+роботи|самост[іi]й[-\s/]*н\w*(?:\s*/\s*|\s+)робот\w*|вільний\s+день)$"
)
DROP_REVIEW_TECHNICAL_RE = re.compile(
    r"(?iu)^(?:classroom\.?|google\s+classroom\.?|гугл\s+клас:?\.?|\[\d{2}\.\d{2}(?:,\s*\d{2}\.\d{2})*\](?:\s*(?:\.|\((?:пр|л|лек|практ|сем|лаб|с|c)\))\s*)?|вкл\.?\s*\d{2}\.\d{2}\.?|(?:іd|id)\s*:\s*\d+(?:\s+\d+)+|понедельник|вторник|среда|четверг|пятница|суббота|воскресенье)$"
)


def partition_rows(rows: list[NormalizedRow], threshold: float) -> tuple[list[NormalizedRow], list[NormalizedRow]]:
    accepted: list[NormalizedRow] = []
    review: list[NormalizedRow] = []
    for row in rows:
        analyze_row_quality(row)
        required_ok = bool(row.day and row.start_time and row.end_time and row.subject)
        hard_fail = any(flag in HARD_FAIL_FLAGS for flag in row.qa_flags)
        if required_ok and row.confidence >= threshold and not hard_fail:
            accepted.append(row)
        else:
            review.append(row)
    return accepted, review


def refine_group_quality(accepted: list[NormalizedRow], review: list[NormalizedRow]) -> tuple[list[NormalizedRow], list[NormalizedRow]]:
    groups: dict[tuple[str, str, str, str, str, str, str, str], list[NormalizedRow]] = defaultdict(list)
    for row in accepted:
        if not row.sheet_name.startswith("pdf"):
            continue
        key = (
            row.source_root_url,
            row.asset_locator,
            row.sheet_name,
            row.day,
            row.start_time,
            row.end_time,
            row.course,
            row.groups,
        )
        groups[key].append(row)

    demoted: set[int] = set()
    for rows in groups.values():
        bare_rows = [
            row
            for row in rows
            if not any(value.strip() for value in (row.teacher, row.room, row.link, row.notes))
        ]
        if len(bare_rows) < 2:
            continue
        if not any(_looks_like_fragment_subject(row.subject) for row in bare_rows):
            continue
        for row in bare_rows:
            row.qa_flags = list(dict.fromkeys([*row.qa_flags, "inconsistent_columns"]))
            row.qa_severity = "fail"
            review.append(row)
            demoted.add(id(row))
    return [row for row in accepted if id(row) not in demoted], review


def analyze_row_quality(row: NormalizedRow) -> NormalizedRow:
    flags = list(dict.fromkeys(row.qa_flags))
    subject = row.subject.strip()
    if not row.day:
        flags.append("missing_day")
    if not (row.start_time and row.end_time):
        flags.append("missing_time")
    if not subject:
        flags.append("missing_subject")
    if subject and looks_like_room_text(subject):
        flags.append("subject_contains_room")
    if subject and looks_like_roomish_subject_text(subject):
        flags.append("subject_contains_room")
    if subject and TRAILING_ROOM_RE.search(subject):
        flags.append("subject_contains_room")
    if subject and looks_like_teacher_text(subject):
        flags.append("subject_contains_teacher")
    if subject and re.search(r"(?u)\b[А-ЯІЇЄҐ][а-яіїєґ'’ʼ-]+\s+[А-ЯІЇЄҐ]\.(?!\s*[А-ЯІЇЄҐ]\.)", subject):
        flags.append("subject_contains_teacher")
    if subject and contains_link_text(subject):
        flags.append("subject_contains_link")
    if subject and looks_like_admin_text(subject):
        flags.append("service_text_subject")
    if subject and len(subject) > 140:
        flags.append("subject_too_long")
    if row.teacher and len(row.teacher) > 180:
        flags.append("teacher_too_long")
    if row.teacher and (LESSON_TEXT_RE.search(row.teacher) or re.search(r"\b\d{3,4}\b", row.teacher) or re.search(r"\b\d{1,2}[.:]\d{2}\b", row.teacher)):
        flags.append("inconsistent_columns")
    if row.room and (looks_like_teacher_text(row.room) or contains_link_text(row.room)):
        flags.append("inconsistent_columns")
    if row.room and len(row.room) > 80:
        flags.append("inconsistent_columns")
    if row.room and re.search(r"(?iu)\b(?:доц|проф|ас|асист|викл)\.?\b", row.room):
        flags.append("inconsistent_columns")
    if subject and looks_like_service_text(subject):
        flags.append("service_text_subject")
    if subject and looks_like_forbidden_subject_text(subject):
        flags.append("service_text_subject")
    if subject and looks_like_garbage_text(subject):
        flags.append("garbage_text")
    if subject and any(token in subject.casefold() for token in ("?pwd=", "?p=", ".us")):
        flags.append("garbage_text")
    if subject and subject.count(" / ") >= 2 and not _looks_like_wrapped_multiline_subject(subject):
        flags.append("inconsistent_columns")
    if subject and re.search(r"\d{2}\.\d{2}\.\d{4}", subject):
        flags.append("inconsistent_columns")
    if subject and ("-----" in subject or "––––" in subject or "лек....." in subject or "практ....." in subject):
        flags.append("inconsistent_columns")
    if row.teacher and ("лек" in row.teacher.casefold() or "практ" in row.teacher.casefold() or row.teacher.count(";") >= 5):
        flags.append("inconsistent_columns")
    if row.notes and len(row.notes) > 240:
        flags.append("inconsistent_columns")
    if row.notes and len(row.notes) > 80 and looks_like_service_text(row.notes):
        flags.append("inconsistent_columns")
    if subject and subject.startswith("="):
        flags.append("garbage_text")
    if subject and subject.isdigit():
        flags.append("garbage_text")
    if subject and not any(character.isalpha() for character in subject):
        flags.append("garbage_text")
    if subject and ABBREVIATED_SUBJECT_RE.fullmatch(subject):
        flags.append("garbage_text")
    if subject and _looks_like_fragment_subject(subject):
        flags.append("inconsistent_columns")
    compact_subject = subject.replace(" ", "")
    if (
        subject
        and " " not in subject
        and re.fullmatch(r"[A-Za-z0-9+/=_-]{8,}", compact_subject)
        and (
            any(character.isdigit() for character in compact_subject)
            or any(character in "+/=_-" for character in compact_subject)
            or (re.search(r"[A-Z]", compact_subject) and re.search(r"[a-z]", compact_subject))
        )
    ):
        flags.append("garbage_text")
    if subject and " " not in subject and re.fullmatch(r"[A-Za-z0-9=._-]{10,}", subject):
        flags.append("garbage_text")
    if _has_implausible_time(row.start_time, row.end_time, subject):
        flags.append("implausible_time")

    severity = "none"
    if any(flag in HARD_FAIL_FLAGS for flag in flags):
        severity = "fail"
    elif flags:
        severity = "warning"

    row.qa_flags = list(dict.fromkeys(flags))
    row.qa_severity = severity
    return row


def sanitize_export_rows(accepted: list[NormalizedRow], review: list[NormalizedRow]) -> tuple[list[NormalizedRow], list[NormalizedRow]]:
    sanitized: list[NormalizedRow] = []
    pending_review = list(review)

    for row in accepted:
        resolved_program = _resolve_program_label(row)
        if not resolved_program:
            row.qa_flags = list(dict.fromkeys([*row.qa_flags, "bad_program_label"]))
            row.qa_severity = "fail"
            pending_review.append(row)
            continue
        if resolved_program != row.program:
            row.program = resolved_program
            row.autofix_actions = list(dict.fromkeys([*row.autofix_actions, "program_label_recovered"]))
        sanitized.append(row)

    final_rows: list[NormalizedRow] = []
    buckets: dict[tuple[str, str], list[NormalizedRow]] = defaultdict(list)
    for row in sanitized:
        buckets[(row.faculty, row.program)].append(row)

    for bucket_rows in buckets.values():
        if (
            _should_force_review_biomed_subject_bucket(bucket_rows)
            or _should_force_review_bucket_by_content(bucket_rows)
            or _should_demote_tiny_program_bucket(bucket_rows)
        ):
            for row in bucket_rows:
                row.qa_flags = list(dict.fromkeys([*row.qa_flags, "bad_program_label"]))
                row.qa_severity = "fail"
                pending_review.append(row)
            continue
        final_rows.extend(bucket_rows)

    filtered_review = [row for row in pending_review if not _should_drop_non_schedule_review_row(row)]
    return final_rows, filtered_review


def _should_force_review_biomed_subject_bucket(rows: list[NormalizedRow]) -> bool:
    if not rows:
        return False
    if rows[0].source_name.casefold() != "biomed-schedule" or len(rows) > 2:
        return False
    program = normalize_service_tokens(rows[0].program).casefold()
    if not program:
        return False
    subjects = {
        normalize_service_tokens(row.subject).casefold()
        for row in rows
        if row.subject.strip()
    }
    if subjects != {"іноземна мова"} and subjects != {"основи підприємництва"}:
        return False
    if any(normalize_service_tokens(row.groups) for row in rows):
        return False
    return all(normalize_service_tokens(row.notes).casefold().startswith(program) for row in rows)


def _should_force_review_bucket_by_content(rows: list[NormalizedRow]) -> bool:
    if not rows:
        return False
    program = normalize_program_candidate(rows[0].program)
    source_name = rows[0].source_name.casefold()
    anchored_program_notes = all(_notes_anchor_program_label(row.notes, program) for row in rows)
    weak_groups = all(not row.groups.strip() for row in rows)
    normalized_subjects = {
        normalize_service_tokens(row.subject)
        for row in rows
        if row.subject.strip()
    }
    if source_name == "history-schedule" and len(rows) <= 3:
        normalized_notes = [
            normalize_service_tokens(row.notes).casefold()
            for row in rows
            if row.notes.strip()
        ]
        program_tokens = {
            token
            for token in re.findall(r"(?u)[а-яіїєґa-z]{5,}", program.casefold())
            if token not in {"курс"}
        }
        if (
            re.match(r"(?u)\d{3}\s+", program)
            and weak_groups
            and normalized_subjects
            and all(subject != program for subject in normalized_subjects)
            and len(normalized_notes) == len(rows)
            and program_tokens
            and all(any(token in note for token in program_tokens) for note in normalized_notes)
        ):
            return True
    if source_name == "biomed-schedule" and len(rows) <= 2:
        if (
            anchored_program_notes
            and weak_groups
            and (
                normalized_subjects == {"іноземна мова"}
                or normalized_subjects == {"основи підприємництва"}
            )
        ):
            return True
    if source_name == "chem-schedule" and len(rows) <= 3:
        if (
            anchored_program_notes
            and weak_groups
            and len(normalized_subjects) == 1
            and re.fullmatch(r"(?u)[A-ZА-ЯІЇЄҐ]{2,5}", next(iter(normalized_subjects)))
        ):
            return True
    return False


def _looks_like_fragment_subject(subject: str) -> bool:
    stripped = subject.strip()
    if not stripped:
        return True
    if re.fullmatch(r"(?u)[А-ЯІЇЄҐA-Z]\.\s*[А-ЯІЇЄҐA-Z]\.?", stripped):
        return True
    if LOWERCASE_DOTTED_VALID_SUBJECT_RE.fullmatch(stripped):
        return False
    if stripped[0].islower():
        return True
    if any(token in stripped.casefold() for token in (".us", "?pwd=", "pwd=", "ідентифікатор", "идентификатор", "конференції", "конференции")):
        return True
    if stripped.count("(") != stripped.count(")"):
        return True
    if stripped.endswith(("-", "/", ":")):
        return True
    if re.fullmatch(r"\([^)]{1,12}\)\s*\d{1,4}", stripped):
        return True
    return False


def _resolve_program_label(row: NormalizedRow) -> str:
    return coalesce_program_label(
        row.program,
        row.groups,
        infer_asset_label_from_locator(row.asset_locator),
        row.sheet_name,
        _program_hint_from_notes(row.notes),
        humanize_source_name(row.source_name),
    )


def _course_as_program_label(course: str) -> str:
    cleaned = normalize_service_tokens(course)
    if not cleaned:
        return ""
    if re.fullmatch(r"\d{1,2}", cleaned):
        return f"{cleaned} курс"
    if re.fullmatch(r"(?iu)\d{1,2}\s*курс", cleaned):
        return cleaned
    return ""


def _program_hint_from_notes(notes: str) -> str:
    cleaned = normalize_service_tokens(notes)
    if not cleaned:
        return ""
    if len(cleaned) < 6:
        return ""
    if contains_link_text(cleaned):
        return ""
    if re.search(r"(?iu)\b(?:ауд|корп)\.?\b", cleaned):
        return ""
    if looks_like_room_text(cleaned):
        return ""
    if looks_like_teacher_text(cleaned):
        return ""
    if looks_like_service_text(cleaned):
        return ""
    if re.search(r"\d{2}\.\d{2}\.\d{4}", cleaned):
        return ""
    if re.search(r"(?u)\b\d{3}\s+\d{3}\s+\d{4}\b", cleaned):
        return ""
    if count_program_codes(cleaned) >= 2:
        return ""
    if re.search(r"(?iu)\b\d{3}\b", cleaned) and any(separator in cleaned for separator in ("/", ";")):
        return ""
    if any(stem in cleaned.casefold() for stem in ("січ", "лют", "берез", "квіт", "трав", "черв", "лип", "серп", "верес", "жовт", "листоп", "груд")) and "тиж" in cleaned.casefold():
        return ""
    if re.search(r"(?iu)\b(?:E\d|0\d{2}|1\d{2})\b.*[,;].*\b(?:E\d|0\d{2}|1\d{2})\b", cleaned):
        return ""
    if looks_like_bad_program_label(cleaned):
        return ""
    return cleaned


def _should_demote_tiny_program_bucket(rows: list[NormalizedRow]) -> bool:
    if not rows:
        return False
    program = normalize_program_candidate(rows[0].program)
    source_name = rows[0].source_name.casefold()
    anchored_program_notes = all(_notes_anchor_program_label(row.notes, program) for row in rows)
    weak_groups = all(not row.groups.strip() for row in rows)
    normalized_subjects = {
        normalize_service_tokens(row.subject)
        for row in rows
        if row.subject.strip()
    }
    if source_name == "history-schedule" and len(rows) <= 3:
        if (
            re.match(r"(?u)\d{3}\s+", program)
            and anchored_program_notes
            and weak_groups
            and normalized_subjects
            and all(subject != program for subject in normalized_subjects)
        ):
            return True
    if source_name == "biomed-schedule" and len(rows) <= 2:
        if (
            anchored_program_notes
            and weak_groups
            and (
                normalized_subjects == {"іноземна мова"}
                or normalized_subjects == {"основи підприємництва"}
            )
        ):
            return True
    if source_name == "biomed-schedule" and len(rows) <= 3:
        normalized_notes = [
            normalize_service_tokens(row.notes)
            for row in rows
            if normalize_service_tokens(row.notes)
        ]
        if (
            program.casefold() == "лаб.діагностика бакалавр"
            and len(normalized_notes) == len(rows)
            and normalized_subjects == {"Ендокринологія з оцінкою результатів досліджень"}
            and any(" ; " in note for note in normalized_notes)
            and all(not _notes_anchor_program_label(row.notes, program) for row in rows)
        ):
            return True
    if source_name == "chem-schedule" and len(rows) <= 3:
        if (
            anchored_program_notes
            and weak_groups
            and len(normalized_subjects) == 1
            and re.fullmatch(r"(?u)[A-ZА-ЯІЇЄҐ]{2,5}", next(iter(normalized_subjects)))
        ):
            return True
    if source_name == "phys-schedule":
        if re.fullmatch(r"(?iu)(?:акад|чл\.-?кор|с\.н\.с|[дк]\.[а-яіїєґ]\.-?[а-яіїєґ]\.[а-яіїєґ]\.?)", program):
            return True
        if re.search(r"(?iu)\b(?:timetable|sem\.)\b", program) or re.search(r"\b20\d{2}\b", program):
            return True
    if len(rows) > 3:
        return False
    if source_name == "iht-schedule" and count_program_codes(program) >= 1:
        return True
    if source_name == "law-schedule" and len(rows) <= 3 and re.fullmatch(r"(?iu)\d+\s*академ\w*", program):
        return True
    if source_name == "econom-schedule" and len(rows) <= 2:
        if re.fullmatch(r"(?u)[А-ЯІЇЄҐ][а-яіїєґ'’ʼ-]+\s+[А-ЯІЇЄҐA-Z]\.(?:[А-ЯІЇЄҐA-Z]\.?)?", program):
            return True
        if (
            anchored_program_notes
            and all(row.subject.strip() and normalize_service_tokens(row.subject) != program for row in rows)
            and re.fullmatch(r"(?u)[А-ЯІЇЄҐ][а-яіїєґ'’ʼ-]+\s+[А-ЯІЇЄҐA-Z]\.(?:[А-ЯІЇЄҐA-Z]\.?)?", program)
        ):
            return True
    if source_name == "history-schedule" and len(rows) <= 1:
        if (
            anchored_program_notes
            and all(row.subject.strip() and normalize_service_tokens(row.subject) != program for row in rows)
        ):
            return True
    if looks_like_bad_program_label(program):
        return True
    if any(pattern.search(program) for pattern in TINY_BAD_PROGRAM_PATTERNS):
        return True
    if _looks_like_tiny_fragmented_program(program):
        return True
    if source_name == "phys-schedule" and len(rows) <= 5:
        safe_markers = ("груп", "курс", "бакалавр", "магістр", "фізик", "оптик", "астроном", "мат-во", "матем", "івт", "наносистем")
        if looks_like_teacher_text(program):
            return True
        if not any(token in program.casefold() for token in safe_markers):
            return True
    if source_name == "phys-schedule" and len(rows) <= 3:
        if any(
            looks_like_room_text(row.subject)
            or looks_like_roomish_subject_text(row.subject)
            or _looks_like_fragment_subject(row.subject)
            for row in rows
        ):
            return True
    if source_name == "biomed-schedule" and len(rows) <= 2:
        if any(_looks_like_fragment_subject(row.subject) for row in rows):
            return True
        if (
            program.casefold() == "іноземна мова"
            and anchored_program_notes
            and all(row.subject.strip() and normalize_service_tokens(row.subject) != program for row in rows)
        ):
            return True
    if source_name == "sociology-schedule" and len(rows) <= 3:
        lowered = program.casefold()
        if lowered in {"english", "англ.мова"}:
            return True
        if re.fullmatch(r"(?iu)\d+\s*маг\b.*", program):
            return True
        if re.fullmatch(r"(?iu)[лпс]-\d+\s*год\.?", program):
            return True
        if _looks_like_uppercase_subject_bucket(program):
            return True
        if looks_like_teacher_text(program) or re.fullmatch(r"(?u)[А-ЯІЇЄҐ][А-ЯІЇЄҐ'’ʼ-]+\s+[А-ЯІЇЄҐ]\.(?:\s*[А-ЯІЇЄҐ]\.?)?", program):
            return True
        if all(
            row.notes.strip() and normalize_service_tokens(row.notes) == program
            for row in rows
        ) and all(
            row.subject.strip() and re.fullmatch(r"(?u)[А-ЯІЇЄҐA-Z][А-ЯІЇЄҐA-Z-]{3,}", row.subject.strip())
            for row in rows
        ):
            return True
        subjects = {normalize_service_tokens(row.subject) for row in rows if row.subject.strip()}
        if (
            subjects
            and len(subjects) == 1
            and next(iter(subjects)) != program
            and weak_groups
            and anchored_program_notes
        ):
            return True
    if source_name == "psy-schedule" and len(rows) == 1:
        note = normalize_service_tokens(rows[0].notes)
        if note and DROP_REVIEW_SERVICE_RE.fullmatch(note):
            return True
    if source_name == "psy-schedule" and len(rows) <= 3:
        normalized_notes = [
            normalize_service_tokens(row.notes)
            for row in rows
            if normalize_service_tokens(row.notes)
        ]
        has_self_study_note = any(DROP_REVIEW_SERVICE_RE.fullmatch(note) for note in normalized_notes)
        if normalized_subjects == {"Базова загальновійськова підготовка"}:
            return True
        if has_self_study_note and "Базова загальновійськова підготовка" in normalized_subjects:
            return True
        if (
            len(rows) == 1
            and normalized_subjects == {"Захист навчальної практики"}
            and rows[0].notes.strip()
            and not _notes_anchor_program_label(rows[0].notes, program)
        ):
            return True
        if len(rows) <= 2 and has_self_study_note and weak_groups:
            return True
    if source_name == "mechmat-schedule" and len(rows) <= 1:
        if "+" in program or LESSON_TEXT_RE.search(program):
            return True
    if len(program) >= 70:
        return True
    generic_labels = {
        "",
        "unknown program",
        "невідома програма",
    }
    generic_labels.update(
        {
            humanize_source_name(rows[0].source_name).casefold(),
            rows[0].source_name.casefold(),
            rows[0].faculty.casefold(),
        }
    )
    return program.casefold() in generic_labels


def _looks_like_tiny_fragmented_program(program: str) -> bool:
    cleaned = normalize_service_tokens(program)
    if not cleaned:
        return True
    if looks_like_teacher_text(cleaned):
        return True
    if cleaned.count("(") != cleaned.count(")"):
        return True
    if re.fullmatch(r"(?iu)[А-ЯІЇЄҐ'’ʼ-]+\s+[А-ЯІЇЄҐ]$", cleaned):
        return True
    if re.fullmatch(r"[A-Za-z0-9]{6,}(?:\.\d+)?", cleaned) and any(ch.isdigit() for ch in cleaned) and any(ch.isupper() for ch in cleaned) and any(ch.islower() for ch in cleaned):
        return True
    if re.fullmatch(r"(?iu)\([^)]{1,12}\)\s*\d.*", cleaned):
        return True
    if re.fullmatch(r"(?iu)\d{1,2}\.\d{1,2}\..*", cleaned):
        return True
    if re.fullmatch(r"(?iu)(?:dr|prof|associate|assistant)\.?\s+[A-ZА-ЯІЇЄҐ][A-Za-zА-ЯІЇЄҐа-яіїєґ'’ʼ.-]+(?:\s+[A-ZА-ЯІЇЄҐ][A-Za-zА-ЯІЇЄҐа-яіїєґ'’ʼ.-]+){0,3}", cleaned):
        return True
    if cleaned[0].islower():
        return True
    if re.fullmatch(r"(?iu)\d{1,2}\s*курс", cleaned):
        return True
    if re.fullmatch(r"(?iu)\d+\s*магістр\w*$", cleaned):
        return True
    if re.fullmatch(r"(?iu)\d+\s*курс\s*\([^)]{2,}\)$", cleaned):
        return True
    if re.search(r"(?iu)\((?:л|пр|лаб|сем)\)$", cleaned):
        return True
    if cleaned.casefold().startswith(("rozklad ", "schedule of classes ")):
        return True
    conjunctions = {"та", "і", "й", "and"}
    words = cleaned.casefold().split()
    if words and (words[0] in conjunctions or words[-1] in conjunctions):
        return True
    if ";" in cleaned:
        return True
    if "+" in cleaned and LESSON_TEXT_RE.search(cleaned):
        return True
    if re.search(r"(?iu)\b\d+\s*год\.?", cleaned):
        return True
    return False


def _notes_anchor_program_label(notes: str, program: str) -> bool:
    cleaned_notes = normalize_service_tokens(notes)
    cleaned_program = normalize_service_tokens(program)
    if not cleaned_notes or not cleaned_program:
        return False
    if cleaned_notes == cleaned_program:
        return True
    if re.fullmatch(rf"{re.escape(cleaned_program)}(?:\s*[\.;,:])+", cleaned_notes):
        return True
    return any(
        re.match(rf"{re.escape(cleaned_program)}\s*{re.escape(suffix)}", cleaned_notes)
        for suffix in (":", ";", ",", ".")
    )


def _should_drop_non_schedule_review_row(row: NormalizedRow) -> bool:
    subject = normalize_service_tokens(row.subject)
    lesson_type = normalize_service_tokens(row.lesson_type)
    notes = normalize_service_tokens(row.notes)
    raw_excerpt = normalize_service_tokens(row.raw_excerpt)

    if subject and _looks_like_drop_review_text(subject):
        return True
    if subject:
        return False

    evidence = [value for value in (lesson_type, notes, raw_excerpt) if value]
    if not evidence:
        return False
    return any(_looks_like_drop_review_text(value) for value in evidence)


def _looks_like_drop_review_text(value: str) -> bool:
    cleaned = normalize_service_tokens(value)
    if not cleaned:
        return False
    if DROP_REVIEW_SERVICE_RE.fullmatch(cleaned):
        return True
    if DROP_REVIEW_TECHNICAL_RE.fullmatch(cleaned):
        return True
    if looks_like_admin_text(cleaned):
        return True
    if looks_like_service_text(cleaned) and not re.search(r"(?iu)\b(?:іспит|залік|захист|екзамен)\b", cleaned):
        return True
    return False


def _looks_like_uppercase_subject_bucket(program: str) -> bool:
    cleaned = normalize_service_tokens(program)
    if not cleaned:
        return False
    letters = [character for character in cleaned if character.isalpha()]
    if not letters:
        return False
    if sum(1 for character in letters if character.isupper()) / len(letters) < 0.9:
        return False
    if re.search(r"(?iu)\((?:с|л|пр)\)?", cleaned):
        return True
    if re.search(r"(?iu)\b\d+\s*год\.?", cleaned):
        return True
    return False


def _looks_like_wrapped_multiline_subject(subject: str) -> bool:
    stripped = subject.strip()
    if stripped.count(" / ") < 2:
        return False
    if (
        contains_link_text(stripped)
        or looks_like_teacher_text(stripped)
        or looks_like_room_text(stripped)
        or looks_like_roomish_subject_text(stripped)
        or TRAILING_ROOM_RE.search(stripped)
        or re.search(r"\d{2}\.\d{2}\.\d{4}", stripped)
    ):
        return False
    segments = [segment.strip() for segment in stripped.split(" / ") if segment.strip()]
    if len(segments) < 3:
        return False
    if any(not any(character.isalpha() for character in segment) for segment in segments):
        return False
    continuation_segments = sum(
        1 for segment in segments[1:] if segment.startswith("(") or segment[0].islower()
    )
    return continuation_segments >= len(segments) - 1 and bool(LESSON_TEXT_RE.search(stripped))


def _has_implausible_time(start_time: str, end_time: str, subject: str = "") -> bool:
    bounds = []
    for value in (start_time, end_time):
        if not value or ":" not in value:
            continue
        try:
            hours, minutes = (int(part) for part in value.split(":", 1))
        except ValueError:
            return True
        bounds.append(hours * 60 + minutes)
    if any(minutes < 7 * 60 or minutes > 22 * 60 + 30 for minutes in bounds):
        return True
    if len(bounds) == 2:
        duration = bounds[1] - bounds[0]
        if duration < 30:
            return True
        if duration > 240:
            if EXTENDED_DURATION_SUBJECT_RE.search(subject) and duration <= 360:
                return False
            if LONG_PRACTICE_SUBJECT_RE.search(subject) and duration <= 360:
                return False
            return True
    return False


def audit_exported_workbooks(exported_files: list[Path], *, output_dir: Path) -> tuple[list[WorkbookQaSummary], Path, Path]:
    summaries = [_audit_single_workbook(path) for path in sorted(exported_files)]
    json_path = output_dir / "qa_report.json"
    xlsx_path = output_dir / "qa_report.xlsx"
    _write_qa_report_json(summaries, json_path)
    _write_qa_report_xlsx(summaries, xlsx_path)
    return summaries, json_path, xlsx_path


def _audit_single_workbook(path: Path) -> WorkbookQaSummary:
    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        return WorkbookQaSummary(
            file_path=path,
            status="fail",
            row_count=0,
            issue_count=1,
            issues=[f"workbook_open_failed: {exc.__class__.__name__}"],
            sheets=[],
        )
    workbook_issues: list[str] = []
    sheet_summaries: list[WorkbookQaSheetSummary] = []
    total_rows = 0

    technical_name = coalesce_label(path.stem, fallback="")
    if not technical_name or looks_like_technical_label(path.stem):
        workbook_issues.append("technical_file_name")
    if looks_like_bad_program_label(path.stem):
        workbook_issues.append("bad_program_label")

    for sheet in workbook.worksheets:
        row_count = 0
        issue_counter: Counter[str] = Counter()
        non_class_rows = 0
        for row_index in range(3, sheet.max_row + 1):
            values = [sheet.cell(row_index, column).value for column in range(1, BODY_COLUMN_COUNT + 1)]
            if not any(value not in ("", None) for value in values):
                continue
            row_count += 1
            week_type, day, start_time, end_time, subject = (str(values[index]).strip() if values[index] is not None else "" for index in range(5))
            teacher = str(values[5]).strip() if values[5] is not None else ""
            if not week_type:
                issue_counter["missing_week_type"] += 1
            if not (day and start_time and end_time and subject):
                issue_counter["missing_required_cells"] += 1
            if subject and "самостій" in subject.casefold():
                non_class_rows += 1
            if subject and looks_like_room_text(subject):
                issue_counter["subject_contains_room"] += 1
            if subject and looks_like_roomish_subject_text(subject):
                issue_counter["subject_contains_room"] += 1
            if subject and looks_like_teacher_text(subject):
                issue_counter["subject_contains_teacher"] += 1
            if subject and contains_link_text(subject):
                issue_counter["subject_contains_link"] += 1
            if subject and looks_like_admin_text(subject):
                issue_counter["service_text_subject"] += 1
            if subject and looks_like_garbage_text(subject):
                issue_counter["garbage_text"] += 1
            if subject and len(subject) > 140:
                issue_counter["subject_too_long"] += 1
            if teacher and len(teacher) > 180:
                issue_counter["teacher_too_long"] += 1
            if _has_implausible_time(start_time, end_time, subject):
                issue_counter["implausible_time"] += 1
        if row_count == 0:
            issue_counter["empty_sheet"] += 1
        elif row_count <= 3 and (looks_like_technical_label(path.stem) or looks_like_bad_program_label(path.stem)):
            issue_counter["suspicious_small_sheet"] += 1

        total_rows += row_count
        sheet_summaries.append(
            WorkbookQaSheetSummary(
                sheet_name=sheet.title,
                row_count=row_count,
                issue_count=sum(issue_counter.values()),
                issues=[name for name, _ in issue_counter.most_common(5)],
            )
        )
        workbook_issues.extend(name for name, _ in issue_counter.items())

    status = "pass"
    if "missing_required_cells" in workbook_issues or "empty_sheet" in workbook_issues or total_rows == 0:
        status = "fail"
    elif workbook_issues:
        status = "warning"

    issue_counter = Counter(workbook_issues)
    return WorkbookQaSummary(
        file_path=path,
        status=status,
        row_count=total_rows,
        issue_count=sum(issue_counter.values()),
        issues=[name for name, _ in issue_counter.most_common(8)],
        sheets=sheet_summaries,
    )


def _write_qa_report_json(summaries: list[WorkbookQaSummary], path: Path) -> None:
    ensure_parent(path)
    payload = [
        {
            "file_path": str(summary.file_path),
            "status": summary.status,
            "row_count": summary.row_count,
            "issue_count": summary.issue_count,
            "issues": summary.issues,
            "sheets": [
                {
                    "sheet_name": sheet.sheet_name,
                    "row_count": sheet.row_count,
                    "issue_count": sheet.issue_count,
                    "issues": sheet.issues,
                }
                for sheet in summary.sheets
            ],
        }
        for summary in summaries
    ]
    path.write_text(
        "[\n" + ",\n".join(json_dumps(item) for item in payload) + "\n]\n",
        encoding="utf-8",
    )


def _write_qa_report_xlsx(summaries: list[WorkbookQaSummary], path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "qa_report"
    headers = ["file", "sheet", "row_count", "issue_count", "status", "top_issues"]
    for column, title in enumerate(headers, start=1):
        sheet.cell(1, column).value = title
    row_index = 2
    for summary in summaries:
        if not summary.sheets:
            sheet.cell(row_index, 1).value = str(summary.file_path)
            sheet.cell(row_index, 3).value = summary.row_count
            sheet.cell(row_index, 4).value = summary.issue_count
            sheet.cell(row_index, 5).value = summary.status
            sheet.cell(row_index, 6).value = ", ".join(summary.issues)
            row_index += 1
            continue
        for qa_sheet in summary.sheets:
            sheet.cell(row_index, 1).value = str(summary.file_path)
            sheet.cell(row_index, 2).value = qa_sheet.sheet_name
            sheet.cell(row_index, 3).value = qa_sheet.row_count
            sheet.cell(row_index, 4).value = qa_sheet.issue_count
            sheet.cell(row_index, 5).value = summary.status
            sheet.cell(row_index, 6).value = ", ".join(qa_sheet.issues or summary.issues)
            row_index += 1
    ensure_parent(path)
    workbook.save(path)
