from __future__ import annotations

from collections import Counter
from io import BytesIO
from pathlib import Path
from typing import Any
import zipfile

from openpyxl import load_workbook

from .utils import clean_numeric_artifact, flatten_multiline, json_dumps


REFERENCE_COLUMNS = [
    "Тиждень",
    "День",
    "Початок",
    "Кінець",
    "Назва предмета",
    "Викладач",
    "Тип заняття",
    "Посилання (якщо є)",
    "Аудиторія (якщо є)",
    "Групи",
    "Курс",
    "Примітки",
]


def audit_manual_reference_zip(zip_path: Path, *, max_rows_per_sheet: int = 200) -> dict[str, Any]:
    """Read manually filled reference workbooks and summarize their layout vocabulary."""
    workbook_count = 0
    sheet_count = 0
    row_count = 0
    canonical_sheets = 0
    week_values: Counter[str] = Counter()
    lesson_type_values: Counter[str] = Counter()
    group_values: Counter[str] = Counter()
    course_values: Counter[str] = Counter()
    title_values: Counter[str] = Counter()
    sheet_name_values: Counter[str] = Counter()
    noncanonical_headers: list[dict[str, str]] = []

    with zipfile.ZipFile(zip_path) as archive:
        entries = [name for name in archive.namelist() if name.lower().endswith(".xlsx") and not Path(name).name.startswith("~$")]
        workbook_count = len(entries)
        for entry in entries:
            workbook = load_workbook(BytesIO(archive.read(entry)), read_only=True, data_only=True)
            for worksheet in workbook.worksheets:
                sheet_count += 1
                sheet_name_values[worksheet.title] += 1
                title = flatten_multiline(worksheet.cell(1, 1).value)
                if title:
                    title_values[title] += 1
                header = [_normalize_reference_header(worksheet.cell(2, column).value) for column in range(1, 13)]
                if header == REFERENCE_COLUMNS:
                    canonical_sheets += 1
                elif len(noncanonical_headers) < 20:
                    noncanonical_headers.append(
                        {
                            "entry": entry,
                            "sheet": worksheet.title,
                            "header": " | ".join(header),
                        }
                    )
                for row in worksheet.iter_rows(min_row=3, max_row=max_rows_per_sheet + 2, values_only=True):
                    values = [flatten_multiline(value) for value in row[:12]]
                    if not any(values):
                        continue
                    row_count += 1
                    if values[0]:
                        week_values[clean_numeric_artifact(values[0])] += 1
                    if values[6]:
                        lesson_type_values[values[6]] += 1
                    if values[9]:
                        group_values[clean_numeric_artifact(values[9])] += 1
                    if values[10]:
                        course_values[clean_numeric_artifact(values[10])] += 1

    return {
        "zip_path": str(zip_path),
        "workbook_count": workbook_count,
        "sheet_count": sheet_count,
        "sampled_data_rows": row_count,
        "canonical_sheets": canonical_sheets,
        "noncanonical_headers": noncanonical_headers,
        "top_titles": _counter_top(title_values),
        "top_sheet_names": _counter_top(sheet_name_values),
        "top_week_values": _counter_top(week_values),
        "top_lesson_type_values": _counter_top(lesson_type_values),
        "top_group_values": _counter_top(group_values),
        "top_course_values": _counter_top(course_values),
    }


def audit_manual_reference_zip_json(zip_path: Path, *, max_rows_per_sheet: int = 200) -> str:
    return json_dumps(audit_manual_reference_zip(zip_path, max_rows_per_sheet=max_rows_per_sheet))


def _normalize_reference_header(value: Any) -> str:
    return flatten_multiline(value).replace("\n", " ")


def _counter_top(counter: Counter[str], *, limit: int = 80) -> list[dict[str, Any]]:
    return [{"value": value, "count": count} for value, count in counter.most_common(limit)]
